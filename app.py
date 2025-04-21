from flask import Flask, g, render_template,request,redirect,url_for,jsonify,send_from_directory,session,make_response,abort,flash,render_template_string
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import extract,func,or_,ForeignKey,text
from sqlalchemy.sql.expression import extract
from openpyxl import load_workbook,Workbook
from datetime import date,datetime
from dateutil.parser import parse
from flask_migrate import Migrate
import logging
from werkzeug.utils import secure_filename
import os
import zipfile
import shutil
import subprocess
import json
from flask import current_app
from google.oauth2 import id_token
from google_auth_oauthlib.flow import Flow
from google.auth.transport import requests
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from googleapiclient.http import MediaFileUpload,MediaIoBaseUpload
import requests as req
import secrets
from dotenv import load_dotenv
import time
import io
import re

load_dotenv()
app = Flask(__name__)
application=app
profile_images_upload_folder = 'static/profile_images'
if not os.path.exists(profile_images_upload_folder):
    os.makedirs(profile_images_upload_folder)

general_upload_folder = 'static/files'
if not os.path.exists(general_upload_folder):
    os.makedirs(general_upload_folder)    

app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///employer.db"
app.config['SQLALCHEMY_BINDS']={'login':"sqlite:///login.db",
                                'delete_user':"sqlite:///delete.db",
                                'resume':"sqlite:///resume.db",
                                'intro':"sqlite:///intro.db",
                                'interview1':"sqlite:///interview1.db",
                                'interview2':"sqlite:///interview2.db",
                                'hr':"sqlite:///hr.db",
                                'dmax_team_leads':"sqlite:///dmax_tl.db",
                                'dmax_interns':"sqlite:///dmax_intrn.db",
                                'dmax_jrqaeng':"sqlite:///dmax_jrqaeg.db",
                                'dmax_qaeng':"sqlite:///dmax_qaeg.db",
                                'dmax_srqaeng':"sqlite:///dmax_srqaeg.db",
                                'week':"sqlite:///week.db",
                                'panels':"sqlite:///panels.db"
                                
                                }
                                  
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config['UPLOAD_FOLDER']=general_upload_folder
app.config['PROFILE_IMAGE_UPLOAD_FOLDER'] = 'static/profile_images'
app.secret_key = os.environ.get('SECRET_KEY')
db = SQLAlchemy(app)
migrate=Migrate(app,db)

os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

CLIENT_ID = os.environ.get('CLIENT_ID')
CLIENT_SECRET=os.environ.get('CLIENT_SECRET')
PUBLIC_KEY=os.environ.get('EMAILJS_PUBLIC_KEY')
SERVICE_ID=os.environ.get('EMAILJS_SERVICE_ID')
TEMPLATE_ID=os.environ.get('EMAILJS_TEMPLATE_ID')

REDIRECT_URI='http://localhost:5000/google_sign_in' # for oauth sign in
REDIRECT_URI_DRIVE='http://localhost:5000/google_drive_callback' #for google_drive
# REDIRECT_URI = 'https://qaoncloud.com/EMP/google_sign_in' # for oauth sign in
# REDIRECT_URI_DRIVE='https://qaoncloud.com/EMP/google_drive_callback' #for google_drive
# REDIRECT_URI = 'https://intern-final-0b4w.onrender.com/google_sign_in'
# REDIRECT_URI_DRIVE='https://intern-final-0b4w.onrender.com/google_drive_callback'


SCOPES = ['openid', 'https://www.googleapis.com/auth/userinfo.email', 'https://www.googleapis.com/auth/userinfo.profile']
drive_scopes=['https://www.googleapis.com/auth/drive']

flow = Flow.from_client_config(
    {
        "web": {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uris": [REDIRECT_URI],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    },
    scopes=SCOPES
)

drive_flow = Flow.from_client_config(
    {
        "web": {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uris": [REDIRECT_URI_DRIVE],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    },
    scopes=drive_scopes
)


location_corrections = {
    "Kollu": "Kollumangudi",
    "TN palayam":"TN Palayam",
    "WDC": "Whitefield DC",
    # add more mappings as needed
}


class Employee(db.Model):
    Sno = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Emp_id = db.Column(db.String(500))
    Name = db.Column(db.String(500))
    Designation = db.Column(db.String(500))
    Department = db.Column(db.String(500))
    Project = db.Column(db.String(500))
    Job_role = db.Column(db.String(500))
    Employment_status = db.Column(db.String(500))
    Joining_date = db.Column(db.String(500))
    Experience = db.Column(db.String(500))
    Location = db.Column(db.String(500))
    Last_promoted = db.Column(db.String(500))
    Comments = db.Column(db.String(500))
    employee_status=db.Column(db.String(500))
      


class Login(db.Model):
    __bind_key__="login"
    id=db.Column(db.Integer,primary_key=True)
    email=db.Column(db.String(500))
    password=db.Column(db.String(200))
    Role=db.Column(db.String(200))
    Name = db.Column(db.String(200))
    MobileNumber = db.Column(db.String(20))
    photo_filename = db.Column(db.String(255))
    
class Delete_user(db.Model):
    __bind_key__="delete_user"
    id=db.Column(db.Integer,primary_key=True) 
    Name=db.Column(db.String(200))
    Date=db.Column(db.String(200))   

class Resume(db.Model):
    __bind_key__="resume"  
    __tablename__ = 'resume' 
    id=db.Column(db.Integer,primary_key=True)
    filename=db.Column(db.String(255), nullable=False)  
    Name=db.Column(db.String(255))
    Email=db.Column(db.String(255))
    Qualification=db.Column(db.String(255))
    Phone=db.Column(db.String(255))
    Location=db.Column(db.String(255))
    Experience=db.Column(db.Integer)
    QA_Lead=db.Column(db.String(255))
    Link=db.Column(db.String(255)) 
    Role=db.Column(db.String(255))
    Expected_CTC=db.Column(db.String(255))
    Actual_CTC=db.Column(db.String(255))
    Notice_period=db.Column(db.String(255))
    Month=db.Column(db.String(255))
    week=db.Column(db.String(255))
    

        
    
class Intro(db.Model):
    __bind_key__="intro"
    __tablename__ = 'intro'
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId = db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))  
    Json_comments = db.Column(db.Text, default='{}')

class Interview1(db.Model):
    __bind_key__="interview1" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))     
    Json_comments = db.Column(db.Text, default='{}')

class Interview2(db.Model):
    __bind_key__="interview2" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))
    Json_comments = db.Column(db.Text, default='{}')    

class Hr(db.Model):
    __bind_key__="hr" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200)) 
    Json_comments = db.Column(db.Text, default='{}')

class Dmax_tl(db.Model):
    __bind_key__="dmax_team_leads"
    id = db.Column(db.Integer, primary_key=True)
    Centre = db.Column(db.String(100))
    EmployeeName = db.Column(db.String(100))
    EmpID = db.Column(db.String(50))
    Designation = db.Column(db.String(100))
    Project = db.Column(db.String(200))
    Month = db.Column(db.String(50))
    Target = db.Column(db.Integer)
    Actual = db.Column(db.Float)
    Production = db.Column(db.Float)  
    Quality = db.Column(db.Float)  
    Attrition = db.Column(db.Float)  
    Skill = db.Column(db.Float)  
    DmaxSharing = db.Column(db.Float)
    OverallDmaxScore = db.Column(db.Float) 

class Dmax_intern(db.Model):
    __bind_key__="dmax_interns"
    id = db.Column(db.Integer, primary_key=True)
    Centre = db.Column(db.String(100))
    EmployeeName = db.Column(db.String(100))
    EmpID = db.Column(db.String(50))
    Designation = db.Column(db.String(100))
    Project = db.Column(db.String(200))
    Month = db.Column(db.String(50))
    Target = db.Column(db.Integer)
    Actual = db.Column(db.Float)
    Production = db.Column(db.Float)  
    Quality = db.Column(db.Float)  
    Attendance = db.Column(db.Float)  
    Skill = db.Column(db.Float)  
    OverallDmaxScore = db.Column(db.Float)

class Dmax_jr_qa_eng(db.Model):
    __bind_key__="dmax_jrqaeng"
    id = db.Column(db.Integer, primary_key=True)
    Centre = db.Column(db.String(100))
    EmployeeName = db.Column(db.String(100))
    EmpID = db.Column(db.String(50))
    Designation = db.Column(db.String(100))
    Project = db.Column(db.String(200))
    Month = db.Column(db.String(50))
    Target = db.Column(db.Integer)
    Actual = db.Column(db.Float)
    Production = db.Column(db.Float)  
    Quality = db.Column(db.Float)  
    Attendance = db.Column(db.Float)  
    Skill = db.Column(db.Float)  
    OverallDmaxScore = db.Column(db.Float)

class Dmax_qa_eng(db.Model):
    __bind_key__="dmax_qaeng"
    id = db.Column(db.Integer, primary_key=True)
    Centre = db.Column(db.String(100))
    EmployeeName = db.Column(db.String(100))
    EmpID = db.Column(db.String(50))
    Designation = db.Column(db.String(100))
    Project = db.Column(db.String(200))
    Month = db.Column(db.String(50))
    Target = db.Column(db.Integer)
    Actual = db.Column(db.Float)
    Production = db.Column(db.Float)  
    Quality = db.Column(db.Float)  
    Attendance = db.Column(db.Float)  
    Skill = db.Column(db.Float)
    New_initiatives= db.Column(db.Float) 
    OverallDmaxScore = db.Column(db.Float)    

class Dmax_sr_qa_eng(db.Model):
    __bind_key__="dmax_srqaeng"
    id = db.Column(db.Integer, primary_key=True)
    Centre = db.Column(db.String(100))
    EmployeeName = db.Column(db.String(100))
    EmpID = db.Column(db.String(50))
    Designation = db.Column(db.String(100))
    Project = db.Column(db.String(200))
    Month = db.Column(db.String(50))
    Target = db.Column(db.Integer)
    Actual = db.Column(db.Float)
    Production = db.Column(db.Float)  
    Quality = db.Column(db.Float)  
    Attendance = db.Column(db.Float)  
    Skill = db.Column(db.Float)
    New_initiatives= db.Column(db.Float) 
    OverallDmaxScore = db.Column(db.Float)    

class Panel(db.Model):
    __bind_key__="panels"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    email = db.Column(db.String(100), nullable=False, unique=True)    

def get_distinct_statistics(): 
    # Querying distinct roles from the Resume table
    distinct_roles = db.session.query(Resume.Role).distinct().all()
    roles_list=[role[0] for role in distinct_roles]
    statistics={}
    statistics_leads={}
    weekly_statistics = {}
    current_month = datetime.now().strftime('%B')
    distinct_leads = db.session.query(Resume.QA_Lead).distinct().all()
    leads_list = [lead[0] for lead in distinct_leads]
    # print(leads_list)

    overall_counts = {
        'intro': {'selected': 0, 'rejected': 0,'hold':0},
        'interview1': {'selected': 0, 'rejected': 0,'hold':0},
        'interview2': {'selected': 0, 'rejected': 0 ,'hold':0},
        'hr': {'selected': 0, 'rejected': 0},
        'grand_total': 0  # To accumulate overall grand total
    }

    for role in roles_list:
        # Get all resume IDs for the current role
        resume_ids = db.session.query(Resume.id).filter(Resume.Role == role).all()        
        resume_ids = [resume_id[0] for resume_id in resume_ids]
        weekly_statistics[role] = {}
        
        for week_label in ['Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5']:
            resume_ids_week = db.session.query(Resume.id).filter(
            Resume.Month == current_month,  # Match the current month
            Resume.week == week_label       # Match the current week
            ).all()
            resume_ids_week = [resume_id[0] for resume_id in resume_ids_week]
            selected_intro_week = db.session.query(Intro).filter(
                Intro.resumeId.in_(resume_ids_week),
                Intro.Status == 'Move to Interview 1'
            ).count()
            rejected_intro_week = db.session.query(Intro).filter(
                Intro.resumeId.in_(resume_ids_week),
                Intro.Status == 'Rejected'
            ).count()
            selected_interview1_week = db.session.query(Interview1).filter(
                Interview1.resumeId.in_(resume_ids_week),
                Interview1.Status == 'Move to Interview 2'
            ).count()
            rejected_interview1_week = db.session.query(Interview1).filter(
                Interview1.resumeId.in_(resume_ids_week),
                Interview1.Status == 'Rejected'
            ).count()
            selected_interview2_week = db.session.query(Interview2).filter(
                Interview2.resumeId.in_(resume_ids_week),
                Interview2.Status == 'Move to HR Round'
            ).count()
            rejected_interview2_week = db.session.query(Interview2).filter(
                Interview2.resumeId.in_(resume_ids_week),
                Interview2.Status == 'Rejected'
            ).count()
            selected_hr_week = db.session.query(Hr).filter(
                Hr.resumeId.in_(resume_ids_week),
                Hr.Status == 'Move to HR Process'
            ).count()
            rejected_hr_week = db.session.query(Hr).filter(
                Hr.resumeId.in_(resume_ids_week),
                Hr.Status == 'Rejected'
            ).count()
            weekly_statistics[role][week_label] = {
            'intro': {
                'selected': selected_intro_week,
                'rejected': rejected_intro_week
                
            },
            'interview1': {
                'selected': selected_interview1_week,
                'rejected': rejected_interview1_week
                
            },
            'interview2': {
                'selected': selected_interview2_week,
                'rejected': rejected_interview2_week
                
            },  
            'hr': {
                'selected': selected_hr_week,
                'rejected': rejected_hr_week
                
            }              
            }

        selected_intro = db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'Move to Interview 1').count()
        rejected_intro= db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'Rejected').count()
        hold_intro= db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'hold').count()

        #interview1
        selected_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'Move to Interview 2').count()
        rejected_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'Rejected').count()
        hold_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'hold').count()

        #interview2
        selected_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'Move to HR Round').count()
        rejected_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'Rejected').count()
        hold_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'hold').count()

        #HR
        selected_hr = db.session.query(Hr).filter(Hr.resumeId.in_(resume_ids), Hr.Status == 'Move to HR Process').count()
        rejected_hr = db.session.query(Hr).filter(Hr.resumeId.in_(resume_ids), Hr.Status == 'Rejected').count()

        statistics[role] = {
             'intro': {
                'selected': selected_intro,
                'rejected': rejected_intro,
                'hold':hold_intro
            },
            'interview1': {
                'selected': selected_interview1,
                'rejected': rejected_interview1,
                'hold':hold_interview1
            },
            'interview2': {
                'selected': selected_interview2,
                'rejected': rejected_interview2,
                'hold': hold_interview2
            },
            'hr': {
                'selected': selected_hr,
                'rejected': rejected_hr
            }
        }
    for lead in leads_list:
        resume_ids = db.session.query(Resume.id).filter(Resume.QA_Lead == lead).all()
        resume_ids = [resume_id[0] for resume_id in resume_ids]

        selected_intro = db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'Move to Interview 1').count()
        rejected_intro = db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'Rejected').count()
        hold_intro = db.session.query(Intro).filter(Intro.resumeId.in_(resume_ids), Intro.Status == 'hold').count()

        selected_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'Move to Interview 2').count()
        rejected_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'Rejected').count()
        hold_interview1 = db.session.query(Interview1).filter(Interview1.resumeId.in_(resume_ids), Interview1.Status == 'hold').count()

        selected_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'Move to HR Round').count()
        rejected_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'Rejected').count()
        hold_interview2 = db.session.query(Interview2).filter(Interview2.resumeId.in_(resume_ids), Interview2.Status == 'hold').count()

        selected_hr = db.session.query(Hr).filter(Hr.resumeId.in_(resume_ids), Hr.Status == 'Move to HR Process').count()
        rejected_hr = db.session.query(Hr).filter(Hr.resumeId.in_(resume_ids), Hr.Status == 'Rejected').count()

        grand_total = (selected_intro + rejected_intro + hold_intro +
                   selected_interview1 + rejected_interview1 +hold_interview1 +
                   selected_interview2 + rejected_interview2 +hold_interview2 +
                   selected_hr + rejected_hr)

        lead_grand_total = (selected_intro + rejected_intro + hold_intro +
                            selected_interview1 + rejected_interview1 + hold_interview1 +
                            selected_interview2 + rejected_interview2 + hold_interview2 +
                            selected_hr + rejected_hr)

        statistics_leads[lead] = {
            'intro': {'selected': selected_intro, 'rejected': rejected_intro,'hold':hold_intro},
            'interview1': {'selected': selected_interview1, 'rejected': rejected_interview1,'hold':hold_interview1},
            'interview2': {'selected': selected_interview2, 'rejected': rejected_interview2,'hold':hold_interview2},
            'hr': {'selected': selected_hr, 'rejected': rejected_hr},
            'grand_total': grand_total
        }
        overall_counts['intro']['selected'] += selected_intro
        overall_counts['intro']['rejected'] += rejected_intro
        overall_counts['intro']['hold'] += hold_intro
        overall_counts['interview1']['selected'] += selected_interview1
        overall_counts['interview1']['rejected'] += rejected_interview1
        overall_counts['interview1']['hold'] += hold_interview1
        overall_counts['interview2']['selected'] += selected_interview2
        overall_counts['interview2']['rejected'] += rejected_interview2
        overall_counts['interview2']['hold'] += hold_interview2
        overall_counts['hr']['selected'] += selected_hr
        overall_counts['hr']['rejected'] += rejected_hr
        overall_counts['grand_total'] += lead_grand_total
           
    return(statistics, statistics_leads,overall_counts,weekly_statistics)    
    


def send_email_with_file_link(file_link,EMAILJS_PUBLIC_KEY,EMAILJS_SERVICE_ID,EMAILJS_TEMPLATE_ID):
    email_data = {
        'service_id': EMAILJS_SERVICE_ID,
        'template_id': EMAILJS_TEMPLATE_ID,
        'user_id': EMAILJS_PUBLIC_KEY,
        'template_params': {
            'file_link': file_link,  # Attach the Drive link to the email
            'to_email': 'varunrram2003@gmail.com@example.com',  # Replace with the recipient's email address
            'subject': 'Resumes Uploaded to Google Drive',
            'message': f'Here is the link to the uploaded file: {file_link}'
        }
    }
    
    # Send the email using the EmailJS API
    response = requests.post(
        'https://api.emailjs.com/api/v1.0/email/send',
        json=email_data,
        headers={'Content-Type': 'application/json'}
    )

    if response.status_code == 200:
        print('Email sent successfully!')
    else:
        print(f'Failed to send email: {response.text}')

def credentials_to_dict(credentials):
    
    return {
        'token': credentials['access_token'],  # Use the public token attribute
        'refresh_token': credentials['refresh_token'],  # Use the public refresh_token attribute
        'scopes': credentials.scopes
    }

# def generate_html_table(resume_details):
#      # Start the HTML string with the table header
#     html = '<table border="1" style="width:100%; border-collapse: collapse;">'
#     html += '<tr><th>Name</th><th>Role</th><th>Experience</th><th>Location</th>'
#     html += '<th>Actual CTC</th><th>Expected CTC</th><th>Phone</th><th>Email</th></tr>'

#     # Create rows using a list comprehension
#     rows = [
#         '<tr>' +
#         ''.join(f'<td>{detail["Name"]}</td><td>{detail["Role"]}</td><td>{detail["Experience"]}</td>'
#                 f'<td>{detail["Location"]}</td><td>{detail["Actual_CTC"]}</td>'
#                 f'<td>{detail["Expected_CTC"]}</td><td>{detail["Phone"]}</td>'
#                 f'<td>{detail["Email"]}</td>') + 
#         '</tr>'
#         for detail in resume_details
#     ]

#     # Join all rows and add to the table
#     html += ''.join(rows)
#     html += '</table>'
#     return html

def create_excel_in_memory(resume_data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Resumes'

    # Adding headers
    headers = ['Name', 'Role', 'Experience', 'Location', 'Actual CTC', 'Expected CTC', 'Phone', 'Email']
    sheet.append(headers)

    # Adding resume data to Excel
    for resume in resume_data:  # <-- Using resume_data here
        sheet.append([
            resume['Name'],           # Adding 'Name'
            resume['Role'],           # Adding 'Role'
            resume['Experience'],     # Adding 'Experience'
            resume['Location'],       # Adding 'Location'
            resume['Actual_CTC'],     # Adding 'Actual CTC'
            resume['Expected_CTC'],   # Adding 'Expected CTC'
            resume['Phone'],          # Adding 'Phone'
            resume['Email']           # Adding 'Email'
        ])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)  # Rewind the BytesIO object for reading
    return output   

def get_month_from_date(date_str):
    """Extracts month from 'yyyy-mm-dd' format date string."""
    return date_str.split('-')[1] if date_str else None               
def extract_data_from_excel():
    wb = load_workbook("employee_data 1.xlsx")
    ws = wb.active
    column_mappings = {
        'Sno': 0,
        'Emp_id': 1,
        'Name': 2,
        'Designation': 3,
        'Department': 4,
        'Project': 5,
        'Job_role': 6,
        'Employment_status': 7,
        'Joining_date': 8,
        'Experience': 9,
        'Location': 10,
        'Last_promoted': 11,
        'Comments': 12
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        
        if not all(cell is None for cell in row):
            Sno = row[column_mappings['Sno']]
            Emp_id = row[column_mappings['Emp_id']]
            Name = row[column_mappings['Name']]
            Designation = row[column_mappings['Designation']]
            Department = row[column_mappings['Department']]
            Project = row[column_mappings['Project']]
            Job_role = row[column_mappings['Job_role']]
            Employment_status = row[column_mappings['Employment_status']]
            Joining_date = row[column_mappings['Joining_date']]
            Experience = row[column_mappings['Experience']]
            formatted_date = None
            
            if isinstance(Joining_date, datetime):
                join_date = Joining_date 
                formatted_date=join_date.strftime("%d-%m-%Y")
                
                day = join_date.day
                month = join_date.month
                year = join_date.year

                current_date = date.today()
                Experience = current_date.year - year

                if (current_date.month, current_date.day) < (month, day):
                    Experience -= 1

                if Experience < 1:
                    Experience = "Less than 1 year"
            else:
                join_date = None
                day = None
                month = None
                year = None
                Experience = None

                 
            Location = row[column_mappings['Location']]
            Last_promoted = row[column_mappings['Last_promoted']]
            Comments = row[column_mappings['Comments']]
            


            existing_data = Employee.query.filter_by(Name=Name).first()
            if not existing_data:
                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                    Department=Department, Project=Project, Job_role=Job_role,
                                    Employment_status=Employment_status, Joining_date=formatted_date,
                                    Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                    Comments=Comments,employee_status="active")
                db.session.add(employee)
    db.session.commit()
    
    db.session.commit()

# def extract_excel_resume():
#     column_mappings = {
#         'Sno': 0,
#         'Emp_id': 1,
#         'Name': 2,
#         'Designation': 3,
#         'Department': 4,
#         'Project': 5,
#         'Job_role': 6,
#         'Employment_status': 7,
#         'Joining_date': 8,
#         'Experience': 9,
#         'Location': 10,
#         'Last_promoted': 11,
#         'Comments': 12
#     }
#     for row in ws.iter_rows(min_row=2, values_only=True):
        
#         if not all(cell is None for cell in row):
#             Sno = row[column_mappings['Sno']]
#             Emp_id = row[column_mappings['Emp_id']]
#             Name = row[column_mappings['Name']]
#             Designation = row[column_mappings['Designation']]
#             Department = row[column_mappings['Department']]
#             Project = row[column_mappings['Project']]
#             Job_role = row[column_mappings['Job_role']]
#             Employment_status = row[column_mappings['Employment_status']]
#             Joining_date = row[column_mappings['Joining_date']]
#             Experience = row[column_mappings['Experience']]
def send_email_via_emailjs(html_content):
    # Use your existing EmailJS configuration
    EMAILJS_USER_ID=os.environ.get('EMAILJS_PUBLIC_KEY')
    EMAILJS_SERVICE_ID=os.environ.get('EMAILJS_SERVICE_ID')
    EMAILJS_TEMPLATE_ID=os.environ.get('EMAILJS_TEMPLATE_ID_TABLE')

    
    

    data = {
        "service_id": EMAILJS_SERVICE_ID,
        "template_id": EMAILJS_TEMPLATE_ID,
        "user_id": EMAILJS_USER_ID,
        "template_params": {
            "html_content": html_content,
            "subject": "Candidate Details",
            "recipient_email": "varunrram2003@gmail.com"  # The recipient's email
        }
    }
    # print(html_content)
    response = req.post('https://api.emailjs.com/api/v1.0/email/send', json=data)

    if response.status_code == 200:
        print("Email sent successfully!")
    else:
        print("Failed to send email:", response.text)
def process_and_insert_data(row, designation):
    
    if designation == 'intern':
        column_mappings = {
                    'S.No': 0,
                    'Centre': 1,
                    'Name of the Employee': 2,
                    'Emp ID': 3,
                    'Designation': 4,
                    'Project (s)': 5,
                    'Month': 6,
                    'Target': 7,
                    'Actual': 8,
                    'Production %': 19,
                    'Quality% (40%)': 20,
                    'Attendance(10%)': 21,
                    'Skill(10%)': 22,
                    'Overall Dmax Score': 23
                }
        centre = row[column_mappings['Centre']]
        name = row[column_mappings['Name of the Employee']]
        emp_id = row[column_mappings['Emp ID']]
        designation = row[column_mappings['Designation']]
        project = row[column_mappings['Project (s)']]
        month = row[column_mappings['Month']]
        target = row[column_mappings['Target']]
        actual = row[column_mappings['Actual']]
        production = format_percentage(row[column_mappings['Production %']])
        quality = format_percentage(row[column_mappings['Quality% (40%)']])
        attendance = format_percentage(row[column_mappings['Attendance(10%)']])
        skill = format_percentage(row[column_mappings['Skill(10%)']])
        overall_dmax_score = format_percentage(row[column_mappings['Overall Dmax Score']])
        
        
        new_employee = Dmax_intern(
                            Centre=centre,
                            EmployeeName=name,
                            EmpID=emp_id,
                            Designation=designation,
                            Project=project,
                            Month=month,
                            Target=target,
                            Actual=actual,
                            Production=production,
                            Quality=quality,
                            Attendance=attendance,
                            Skill=skill,
                            OverallDmaxScore=overall_dmax_score
                        )
        db.session.add(new_employee)
        records_added = True
    if designation == 'jr. qa engineer' or designation =="qa automation engineer":
        column_mappings = {
                    'S.No': 0,
                    'Centre': 1,
                    'Name of the Employee': 2,
                    'Emp ID': 3,
                    'Designation': 4,
                    'Project (s)': 5,
                    'Month': 6,
                    'Target': 7,
                    'Actual': 8,
                    'Production %':19,
                    'Quality% (40%)': 20,
                    'Attendance(10%)':21,
                    'Skill(10%)':22,
                    'Overall Dmax Score':23
                }
        centre = row[column_mappings['Centre']]
        name = row[column_mappings['Name of the Employee']]
        emp_id = row[column_mappings['Emp ID']]
        designation = row[column_mappings['Designation']]
        project = row[column_mappings['Project (s)']]
        month = row[column_mappings['Month']]
        target = row[column_mappings['Target']]
        actual = row[column_mappings['Actual']]
        production = format_percentage(row[column_mappings['Production %']])
        quality = format_percentage(row[column_mappings['Quality% (40%)']])
        attendance = format_percentage(row[column_mappings['Attendance(10%)']])
        skill = format_percentage(row[column_mappings['Skill(10%)']])
        overall_dmax_score = format_percentage(row[column_mappings['Overall Dmax Score']])
        new_employee = Dmax_jr_qa_eng(
                            Centre=centre,
                            EmployeeName=name,
                            EmpID=emp_id,
                            Designation=designation,
                            Project=project,
                            Month=month,
                            Target=target,
                            Actual=actual,
                            Production=production,
                            Quality=quality,
                            Attendance=attendance,
                            Skill=skill,
                            OverallDmaxScore=overall_dmax_score
                        )
        db.session.add(new_employee)
        records_added = True


    if designation == 'sr. qa engineer':
        
        column_mappings = {
                    'S.No': 0,
                    'Centre': 1,
                    'Name of the Employee': 2,
                    'Emp ID': 3,
                    'Designation': 4,
                    'Project (s)': 5,
                    'Month': 6,
                    'Target': 7,
                    'Actual': 8,
                    'Production %': 22,
                    'Quality% (40%)': 23,
                    'Attendance(10%)': 24,
                    'Skill(10%)': 25,
                    'New initiatives (10%)':26,
                    'Overall Dmax Score': 27
                }

        centre = row[column_mappings['Centre']]
        name = row[column_mappings['Name of the Employee']]
        emp_id = row[column_mappings['Emp ID']]
        designation = row[column_mappings['Designation']]
        project = row[column_mappings['Project (s)']]
        month = row[column_mappings['Month']]
        target = row[column_mappings['Target']]
        actual = row[column_mappings['Actual']]
        production = format_percentage(row[column_mappings['Production %']])
        quality = format_percentage(row[column_mappings['Quality% (40%)']])
        attendance = format_percentage(row[column_mappings['Attendance(10%)']])
        skill = format_percentage(row[column_mappings['Skill(10%)']])
        initiatives=format_percentage(row[column_mappings['New initiatives (10%)']])
        overall_dmax_score = format_percentage(row[column_mappings['Overall Dmax Score']])
        new_employee =Dmax_sr_qa_eng(
                            Centre=centre,
                            EmployeeName=name,
                            EmpID=emp_id,
                            Designation=designation,
                            Project=project,
                            Month=month,
                            Target=target,
                            Actual=actual,
                            Production=production,
                            Quality=quality,
                            Attendance=attendance,
                            Skill=skill,
                            New_initiatives=initiatives,
                            OverallDmaxScore=overall_dmax_score
                        )
        db.session.add(new_employee)
        records_added = True

    if designation == 'qa engineer':
        column_mappings = {
                    'S.No': 0,
                    'Centre': 1,
                    'Name of the Employee': 2,
                    'Emp ID': 3,
                    'Designation': 4,
                    'Project (s)': 5,
                    'Month': 6,
                    'Target': 7,
                    'Actual': 8,
                    'Production %': 22,
                    'Quality% (40%)': 23,
                    'Attendance(10%)': 24,
                    'Skill(10%)': 25,
                    'New initiatives (10%)':26,
                    'Overall Dmax Score': 27
                }

        centre = row[column_mappings['Centre']]
        name = row[column_mappings['Name of the Employee']]
        emp_id = row[column_mappings['Emp ID']]
        designation = row[column_mappings['Designation']]
        project = row[column_mappings['Project (s)']]
        month = row[column_mappings['Month']]
        target = row[column_mappings['Target']]
        actual = row[column_mappings['Actual']]
        production = format_percentage(row[column_mappings['Production %']])
        quality = format_percentage(row[column_mappings['Quality% (40%)']])
        attendance = format_percentage(row[column_mappings['Attendance(10%)']])
        skill = format_percentage(row[column_mappings['Skill(10%)']])
        initiatives=format_percentage(row[column_mappings['New initiatives (10%)']])
        overall_dmax_score = format_percentage(row[column_mappings['Overall Dmax Score']])
        new_employee =Dmax_qa_eng(
                            Centre=centre,
                            EmployeeName=name,
                            EmpID=emp_id,
                            Designation=designation,
                            Project=project,
                            Month=month,
                            Target=target,
                            Actual=actual,
                            Production=production,
                            Quality=quality,
                            Attendance=attendance,
                            Skill=skill,
                            New_initiatives=initiatives,
                            OverallDmaxScore=overall_dmax_score
                        )
        db.session.add(new_employee)
        records_added = True

    if designation == 'adm' or designation== 'qa lead':
            column_mappings = {
                        'S.No': 0,
                        'Centre': 1,
                        'Name of the Employee': 2,
                        'Emp ID': 3,
                        'Designation': 4,
                        'Project (s)': 5,
                        'Month': 6,
                        'Target': 7,
                        'Actual': 8,
                        'Production (40%)': 24,
                        'Attrition(10%)': 25,
                        'Quality% (40%)': 26,
                        'Skill(10%)': 27,
                        'Dmax Sharing': 28,
                        'Overall Dmax Score': 29
                    }          
            
            centre = row[column_mappings['Centre']]
            name = row[column_mappings['Name of the Employee']]
            emp_id = row[column_mappings['Emp ID']]
            designation = row[column_mappings['Designation']]
            project = row[column_mappings['Project (s)']]
            month = row[column_mappings['Month']]
            target = row[column_mappings['Target']]
            actual = row[column_mappings['Actual']]
            production = row[column_mappings['Production (40%)']]
            quality = row[column_mappings['Quality% (40%)']]
            attrition = row[column_mappings['Attrition(10%)']]
            skill = row[column_mappings['Skill(10%)']]
            dmax_sharing = row[column_mappings['Dmax Sharing']]
            overall_dmax_score = row[column_mappings['Overall Dmax Score']]   
                    
            production = format_percentage(production)
            quality = format_percentage(quality)
            attrition = format_percentage(attrition)
            skill = format_percentage(skill)
            dmax_sharing=format_percentage(dmax_sharing)
            overall_dmax_score=format_percentage(overall_dmax_score)
            new_employee = Dmax_tl(
                                Centre=centre,
                                EmployeeName=name,
                                EmpID=emp_id,
                                Designation=designation,
                                Project=project,
                                Month=month,
                                Target=target,
                                Actual=actual,
                                Production=production,
                                Quality=quality,
                                Attrition=attrition,
                                Skill=skill,
                                DmaxSharing=dmax_sharing,
                                OverallDmaxScore=overall_dmax_score
                                
                            )
            db.session.add(new_employee)
            records_added = True
    db.session.commit()        
    return records_added        



    




def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1] in ["xlsx","csv"]

def allowed_files(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ["pdf","doc","docx"]

def format_percentage(value):
    if isinstance(value, str):
        value = value.replace('%', '').strip()
        try:
            value=float(value)
        except ValueError:
            print(f"Error converting string value: {value}")
            return None
    elif isinstance(value, (int, float)):
        value=float(value)
    else:
        print(f"Unexpected type: {type(value)}")
        return None
    return round(value * 100, 2)

def map_role_based_on_experience(experience):
    # Convert months to years, assuming experience can be a float like 0.5 for 6 months
    if experience:
        experience_in_years = float(experience)
    
        # Define the role based on the years of experience
        if experience_in_years >= 6:
            return "QA Lead"
        elif 4 <= experience_in_years < 6:
            return "Sr. QA Engineer"
        elif 2 <= experience_in_years < 4:
            return "QA Engineer"
        elif 0.5 <= experience_in_years < 2:
            return "Jr. QA Engineer"
        else:
            return "QA Intern"

    return None


def dashboard_function():
    total_employees=Employee.query.count()
    active_employees = Employee.query.filter_by(employee_status='active').count()
    resigned_employees=Employee.query.filter_by(employee_status='resigned').count()
    total_resumes=Resume.query.count()
    hr_selected=Hr.query.filter_by(Status='Move to HR Process').count()
    current_year = datetime.now().year
    last_year = current_year - 1
    
    
    current_year_count = Employee.query.filter(Employee.Joining_date.like(f'%-{current_year}')).count()
    employment_status_counts = {}
    project_status_counts = {}
    status = Employee.query.with_entities(Employee.Employment_status).distinct()
    project = Employee.query.with_entities(Employee.Project).distinct()
    for stat in status:
        count = Employee.query.filter_by(Employment_status=stat.Employment_status).count()
        employment_status_counts[stat.Employment_status] = count

    for proj in project:
        count = Employee.query.filter_by(Project=proj.Project).count()
        project_status_counts[proj.Project] = count
    last_year_count = Employee.query.filter(Employee.Joining_date.like(f'%-{last_year}')).count()
    Chennai_employees_count=Employee.query.filter_by(Location='Chennai').count()
    kollu_employees_count=Employee.query.filter_by(Location='Kollumangudi').count()
    kaup_employees_count=Employee.query.filter_by(Location='Kaup').count()
    tn_palyam_count=Employee.query.filter_by(Location='TN Palayam').count()
    return(total_employees,active_employees,resigned_employees,total_resumes,hr_selected,current_year_count, last_year_count,Chennai_employees_count,kollu_employees_count,kaup_employees_count,tn_palyam_count,employment_status_counts, project_status_counts)
@app.context_processor
def inject_total_employees():
    total_employees,active_employees,resigned_employees,total_resumes,hr_selected,current_year_count, last_year_count,Chennai_employees_count,kollu_employees_count,kaup_employees_count,tn_palyam_count,employment_status_counts, project_status_counts=dashboard_function()
    statistics,statistics_leads,overall_counts,weekly_statistics=get_distinct_statistics()
    return dict(total_employees=total_employees,active_employees=active_employees,resigned_employees=resigned_employees,total_resumes=total_resumes,hr_selected=hr_selected,current_year_count=current_year_count, last_year_count=last_year_count,Chennai_employees_count=Chennai_employees_count,kollu_employees_count=kollu_employees_count,kaup_employees_count=kaup_employees_count,tn_palyam_count=tn_palyam_count,employment_status_counts=employment_status_counts,
                project_status_counts=project_status_counts,statistics=statistics,statistics_leads=statistics_leads,overall_counts=overall_counts,weekly_statistics=weekly_statistics)
@app.before_request
def load_user():
    
    email = session.get('email')
    print(f"Email from session: {email}") 
    if email:
        
        user = Login.query.filter_by(email=email).first()
        
        if user:
            
            g.user_role = user.Role
            g.user_name = user.Name
        else:
            print("No user found")    
    

@app.route("/",methods=["GET","POST"])
def signPage():
    correct_user=None
    error_message=None
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        
        correct_user=Login.query.filter_by(email=email).first()
        if correct_user:
            if correct_user.password==password:
                session['email'] = email 
                if correct_user.Role=="admin":
                    
                    return redirect(url_for("dashBoard"))
                else:
                    return redirect(url_for("employee"))        
            else:
                correct_user=None
                error_message="invalid login credentials"
        if correct_user == None:
                error_message="invalid login credentials"      
    return render_template("sign.html",error_message=error_message)   
@app.route("/dashboard")
def dashBoard():
    data = {"Intro Selected": 10, "Intro Rejected": 5}
    chart_data = {
        "labels": list(data.keys()),
        "values": list(data.values())  # ✅ note the () after values
    }
    # status=Employee.query.with_entities(Employee.Employment_status).distinct()
    email=session.get('email')
    project_filter = request.args.get('project', '').strip()
    designation_filter= request.args.get('designation', '').strip()
    location_filter= request.args.get('location', '').strip()
    week_filter=request.args.get('week', '').strip()
    month_filter=request.args.get('month', '').strip()
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    if not email:
        return redirect(url_for('login'))
    
    user=Login.query.filter_by(email=email).first()
    
        
    if not user:
        return redirect(url_for('login'))  

    user_role=user.Role
    base_query = Employee.query
    if project_filter:
        base_query = base_query.filter(func.trim(Employee.Project) == project_filter)
    if designation_filter:
        base_query = base_query.filter(func.trim(Employee.Designation) == designation_filter) 
    if location_filter:
        corrected_location_filter = location_corrections.get(location_filter, location_filter)
        base_query = base_query.filter(func.trim(Employee.Location) == corrected_location_filter)        

    # 🔹 Get all distinct locations
    trimmed_locations = db.session.query(func.trim(Employee.Location)).distinct()
    projects = db.session.query(func.trim(Employee.Project)).distinct()
    Employment_status = db.session.query(func.trim(Employee.Employment_status)).distinct()
    employee_status = db.session.query(func.trim(Employee.employee_status)).distinct()
    # Count employees per column
    location_counts = {}
    projects_counts = {}
    Employment_status_counts = {}
    employee_status_counts={}

    

    ## for loop for storing the column counts

    for loc_tuple in trimmed_locations:
        loc = loc_tuple[0]
        count = base_query.filter(func.trim(Employee.Location) == loc).count()
        location_counts[loc] = count

    for proj_tuple in projects:
        proj= proj_tuple[0]
        count=base_query.filter(func.trim(Employee.Project)==proj).count()
        projects_counts[proj]=count

    for Emp_status in Employment_status:
        status=Emp_status[0]
        count=base_query.filter(func.trim(Employee.Employment_status)==status).count()
        Employment_status_counts[status]=count    

    for emp_status in employee_status:
        status=emp_status[0]
        count=base_query.filter(func.trim(Employee.employee_status)==status).count()
        employee_status_counts[status]=count

    total_employees_from_base = base_query.count()
    employee_status_counts["Total"] = total_employees_from_base      

    ## Resume base query

    base_query_resume = db.session.query(Resume)  # No filters for now
    
    if hasattr(g, 'user_name') and getattr(g, 'user_role', None) == 'interviewer':
        base_query_resume = base_query_resume.filter(
            func.lower(func.trim(Resume.QA_Lead)) == g.user_name.lower()
        )

    if week_filter:
        base_query_resume = base_query_resume.filter(
            func.lower(func.trim(Resume.week)) == week_filter.lower()
        )

    if month_filter:
        base_query_resume = base_query_resume.filter(
            func.lower(func.trim(Resume.Month)) == month_filter.lower()
        )         
    ## Overall counts

    # Get all resume IDs
    resume_ids = [res.id for res in base_query_resume.all()]

    # A helper function that counts distinct statuses from any interview stage
    def get_status_counts(model, resume_ids):
        from collections import defaultdict
        counts = defaultdict(int)

        if not resume_ids:
            return counts

        # Get all unique statuses for this stage
        distinct_statuses = db.session.query(func.trim(model.Status)).filter(
            model.resumeId.in_(resume_ids)
        ).distinct().all()

        for status_tuple in distinct_statuses:
            status = status_tuple[0]
            count = db.session.query(model).filter(
                model.resumeId.in_(resume_ids),
                func.trim(model.Status) == status
            ).count()
            counts[status] = count

        # Add total count for the stage
        counts["Total"] = db.session.query(model).filter(model.resumeId.in_(resume_ids)).count()

        return counts

    # Get counts for all stages
    intro_counts = get_status_counts(Intro, resume_ids)
    interview1_counts = get_status_counts(Interview1, resume_ids)
    interview2_counts = get_status_counts(Interview2, resume_ids)
    hr_counts = get_status_counts(Hr, resume_ids)
    grand_total = (
        intro_counts.get("Total", 0) +
        interview1_counts.get("Total", 0) +
        interview2_counts.get("Total", 0) +
        hr_counts.get("Total", 0)
    )

    overall_counts = {
        "grand_total": grand_total
    }
    

    ##role specific counts
    roles = [r[0] for r in base_query_resume.with_entities(Resume.Role).distinct()]
    role_intro_stats = {}
    role_interview1_stats = {}
    role_interview2_stats={}
    role_hr_stats={}

    for role in roles:
        resume_ids = [r.id for r in base_query_resume.filter(Resume.Role == role)]
        role_intro_stats[role] = get_status_counts(Intro, resume_ids) 
        role_interview1_stats[role] = get_status_counts(Interview1, resume_ids) 
        role_interview2_stats[role] = get_status_counts(Interview2, resume_ids)
        role_hr_stats[role] = get_status_counts(Hr, resume_ids)

      
    
    role_statistics = {}

    for role in role_intro_stats:
        intro_stats = role_intro_stats.get(role, {})
        interview1_stats = role_interview1_stats.get(role, {})
        interview2_stats = role_interview2_stats.get(role, {})
        hr_stats = role_hr_stats.get(role, {})

        role_statistics[role] = {
            'intro': {
                'selected': intro_stats.get("Move to Interview 1", 0),
                'rejected': intro_stats.get("Rejected", 0),
                'hold': intro_stats.get("On Hold", 0)
            },
            'interview1': {
                'selected': interview1_stats.get("Move to Interview 2", 0),
                'rejected': interview1_stats.get("Rejected", 0),
                'hold': interview1_stats.get("On Hold", 0)
            },
            'interview2': {
                'selected': interview2_stats.get("Move to HR", 0),
                'rejected': interview2_stats.get("Rejected", 0),
                'hold': interview2_stats.get("On Hold", 0)
            },
            'hr': {
                'selected': hr_stats.get("Move to HR Process", 0),
                'rejected': hr_stats.get("Rejected", 0),
                'hold': hr_stats.get("On Hold", 0)
            }
        }   
    ## LEAD STATISTICS
    distinct_leads = base_query_resume.filter(Resume.QA_Lead.isnot(None)).with_entities(Resume.QA_Lead).distinct().all()
    leads = [lead[0] for lead in distinct_leads]
    lead_intro_stats = {}
    lead_interview1_stats = {}
    lead_interview2_stats = {}
    lead_hr_stats = {}

    # Step 3: Loop through each lead and calculate stats
    for lead in leads:
        resume_ids = [r.id for r in base_query_resume.filter(Resume.QA_Lead == lead)]
        lead_intro_stats[lead] = get_status_counts(Intro, resume_ids)
        lead_interview1_stats[lead] = get_status_counts(Interview1, resume_ids)
        lead_interview2_stats[lead] = get_status_counts(Interview2, resume_ids)
        lead_hr_stats[lead] = get_status_counts(Hr, resume_ids)

    # Step 4: Build the final lead_statistics structure
    lead_statistics = {}

    for lead in lead_intro_stats:
        intro_stats = lead_intro_stats.get(lead, {})
        interview1_stats = lead_interview1_stats.get(lead, {})
        interview2_stats = lead_interview2_stats.get(lead, {})
        hr_stats = lead_hr_stats.get(lead, {})

        lead_statistics[lead] = {
            'intro': {
                'selected': intro_stats.get("Move to Interview 1", 0),
                'rejected': intro_stats.get("Rejected", 0),
                'hold': intro_stats.get("On Hold", 0)
            },
            'interview1': {
                'selected': interview1_stats.get("Move to Interview 2", 0),
                'rejected': interview1_stats.get("Rejected", 0),
                'hold': interview1_stats.get("On Hold", 0)
            },
            'interview2': {
                'selected': interview2_stats.get("Move to HR", 0),
                'rejected': interview2_stats.get("Rejected", 0),
                'hold': interview2_stats.get("On Hold", 0)
            },
            'hr': {
                'selected': hr_stats.get("Move to HR Process", 0),
                'rejected': hr_stats.get("Rejected", 0),
                'hold': hr_stats.get("On Hold", 0)
            }
        }
     
    raw_data = {
        "Intro Round Selected": intro_counts.get("Move to Interview 1", 0),
        "Intro Round Rejected": intro_counts.get("Rejected", 0),
        "Intro Round Hold": intro_counts.get("On Hold", 0),
        "L1 Selected": interview1_counts.get("Move to Interview 2", 0),
        "L1 Rejected": interview1_counts.get("Rejected", 0),
        "L1 Hold": interview1_counts.get("On Hold", 0),
        "L2 Selected": interview2_counts.get("Move to HR Round", 0),
        "L2 Rejected": interview2_counts.get("Rejected", 0),
        "L2 Hold": interview2_counts.get("On Hold", 0),
        "HR Selected": hr_counts.get("Move to HR Process", 0),
        "HR Rejected": hr_counts.get("Rejected", 0),
        "HR Hold": hr_counts.get("On Hold", 0)
    }
    total_selected = raw_data.get('Intro Round Selected', 0) + raw_data.get('L1 Selected', 0)  + raw_data.get('L2 Selected', 0) + raw_data.get('HR Selected', 0)
    total_rejected = raw_data.get('Intro Round Rejected', 0)
    total_hold = raw_data.get('L1 Hold', 0)
    
    # Remove all zero values
    chart_labels = [label for label, value in raw_data.items() if value != 0]
    chart_values = [value for value in raw_data.values() if value != 0]
    selected_hover = []
    rejected_hover = []
    hold_hover = []
    total_selected = 0
    lead_breakdowns = {
        'Selected': {},
        'Rejected': {},
        'Hold': {}
    }

    for lead, stats in lead_statistics.items():
        selected_total = stats['intro']['selected'] + stats['interview1']['selected'] + stats['interview2']['selected'] + stats['hr']['selected']
        rejected_total = stats['intro']['rejected'] + stats['interview1']['rejected'] + stats['interview2']['rejected'] + stats['hr']['rejected']
        hold_total = stats['intro']['hold'] + stats['interview1']['hold'] + stats['interview2']['hold'] + stats['hr']['hold']

        selected_hover.append(f"{lead}: {selected_total}")
        rejected_hover.append(f"{lead}: {rejected_total}")
        hold_hover.append(f"{lead}: {hold_total}")

        lead_breakdowns['Selected'][lead] = selected_total
        lead_breakdowns['Rejected'][lead] = rejected_total
        lead_breakdowns['Hold'][lead] = hold_total

    hover_texts = [
        "Selected:<br>" + "<br>".join(selected_hover),
        "Rejected:<br>" + "<br>".join(rejected_hover),
        "Hold:<br>" + "<br>".join(hold_hover)
    ]
    ## Pie charts for Lead wise charts
    labels = ["Interview Results"]
    parents = [""]
    values = [0]  # we'll update this total at the end

    rounds = ['intro', 'interview1', 'interview2', 'hr']
    round_labels = {
        'intro': 'Intro Round',
        'interview1': 'Interview 1',
        'interview2': 'Interview 2',
        'hr': 'HR Round'
    }
    status_labels = {'selected': 'Selected', 'rejected': 'Rejected', 'hold': 'Hold'}

    round_totals = {r: 0 for r in rounds}

    # Step 1: Add top-level rounds
    for r in rounds:
        labels.append(round_labels[r])
        parents.append("Interview Results")
        values.append(0)  # we'll update these later

    # Step 2: Add per-lead stats under each round
    for lead, stats in lead_statistics.items():
        for r_idx, r in enumerate(rounds):
            lead_total = sum(stats[r].values())
            if lead_total == 0:
                continue

            lead_label = f"{lead} ({round_labels[r]})"
            round_label = round_labels[r]

            labels.append(lead_label)
            parents.append(round_label)
            values.append(lead_total)
            round_totals[r] += lead_total

            for status_key, status_label in status_labels.items():
                count = stats[r].get(status_key, 0)
                if count == 0:
                    continue
                labels.append(f"{lead} - {status_label} ({round_labels[r]})")
                parents.append(lead_label)
                values.append(count)

    # Step 3: Update total and round values
    total = sum(round_totals.values())
    values[0] = total  # Interview Results
    for i, r in enumerate(rounds):
        values[i + 1] = round_totals[r]

    # Output JavaScript arrays
    print("labels =", labels)
    print("parents =", parents)
    print("values =", values)
    
    ## Pie charts for Role wise charts

    role_labels = ["Interview Results"]
    role_parents = [""]
    role_values = [0]  # will be updated at the end

    rounds = ['intro', 'interview1', 'interview2', 'hr']
    round_labels = {
        'intro': 'Intro Round',
        'interview1': 'Interview 1',
        'interview2': 'Interview 2',
        'hr': 'HR Round'
    }
    status_labels = {'selected': 'Selected', 'rejected': 'Rejected', 'hold': 'Hold'}

    role_totals = {}  # To store total per role

    # Optional: Clean up None keys to a string
    role_statistics = {role if role else "None": stats for role, stats in role_statistics.items()}

    # Step 1: Add each role as top-level under "Interview Results"
    for role, stats in role_statistics.items():
        total_for_role = 0

        # Count all rounds' totals
        for r in rounds:
            total_for_role += sum(stats.get(r, {}).values())

        # Add role node (even if 0, so we see "None" roles too)
        role_labels.append(role)
        role_parents.append("Interview Results")
        role_values.append(total_for_role)
        role_totals[role] = total_for_role

        # Step 2: Add rounds under each role
        for r in rounds:
            round_data = stats.get(r, {})
            round_total = sum(round_data.values())

            round_label = f"{round_labels[r]} ({role})"
            role_labels.append(round_label)
            role_parents.append(role)
            role_values.append(round_total)

            # Step 3: Add status under each round
            for status_key, status_label in status_labels.items():
                count = round_data.get(status_key, 0)
                if count == 0:
                    continue

                status_label_full = f"{status_label} ({round_labels[r]})"
                role_labels.append(status_label_full)
                role_parents.append(round_label)
                role_values.append(count)

    # Step 4: Update top-level total
    role_values[0] = sum(role_totals.values())
    return render_template("dashboard.html",user_role=user_role,location_counts=location_counts,projects_counts=projects_counts,Employment_status_counts=Employment_status_counts,employee_status_counts=employee_status_counts,intro_counts=intro_counts,interview1_counts=interview1_counts,interview2_counts=interview2_counts,hr_counts=hr_counts,role_statistics=role_statistics,project=project_filter,designation=designation_filter,location=location_filter,months=months,week_filter=week_filter,month_filter=month_filter,lead_statistics=lead_statistics,overall_counts=overall_counts,chart_labels=chart_labels,chart_values=chart_values,hover_texts=hover_texts,total_selected=total_selected,total_rejected=total_rejected,total_hold=total_hold,lead_breakdowns=lead_breakdowns,labels=json.dumps(labels), parents=json.dumps(parents), values=json.dumps(values),role_labels=json.dumps(role_labels),
        role_parents=json.dumps(role_parents),
        role_values=json.dumps(role_values)) 
# @app.route("/home",methods=["GET","POST"])
# def Home():
#     default_page_size = 20
#     page_size_options = [20, 30, 40]

#     # Handle form submission for page size change
#     if request.method == "POST":
#         selected_page_size = int(request.form["page_size"])
#         session['page_size'] = selected_page_size  # Update session variable
#     else:
#         selected_page_size = session.get('page_size', default_page_size)    
#     search_query = request.args.get('search', '')
    
#     # Get data based on current page and retrieved/default page size
#     page = request.args.get('page', 1, type=int)
#     if search_query:
#         total_items = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).count()
#         data = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).paginate(page=page, per_page=selected_page_size)
#     else:
#         total_items = Employee.query.count()
#         data = Employee.query.paginate(page=page, per_page=selected_page_size)

#     # Retrieve selected page size from session (or default)
#     selected_page_size = session.get('page_size', default_page_size)

#     if selected_page_size != default_page_size:
#         new_page_count = total_items // selected_page_size
#         if total_items % selected_page_size > 0:
#             new_page_count += 1  # Account for partial last page
#         page = min(page, new_page_count)
#     start_index = (page - 1) * selected_page_size
#     data = Employee.query.paginate(page=page, per_page=selected_page_size)

#     total_pages=data.pages

#     return render_template("index.html", data=data,
#                                    page_size_options=page_size_options,
#                                    selected_page_size=selected_page_size,total_items=total_items,total_pages=total_pages,start_index=start_index,search_query=search_query)
@app.route("/home", methods=["GET", "POST"])
def Home():
    default_page_size = 20
    page_size_options = [20, 30, 40,'All']
    
    # Handle form submission for page size change
    if request.method == "POST":
        selected_page_size = request.form.get("page_size", default_page_size)
        if selected_page_size != 'All':
            selected_page_size = int(selected_page_size)
        session['page_size'] = selected_page_size
        # Redirect to the same page with updated page size to avoid form resubmission issues
        return redirect(url_for('Home', page=1, search=request.args.get('search', '')))
    else:
        selected_page_size = session.get('page_size', default_page_size)

    project = request.args.get('project')
    designation = request.args.get('designation')
    employment_status = request.args.get('employment_status')
    status = request.args.get('status')
    
    loc = request.args.get('location')
    if loc=="TN palayam":
        loc="TN Palayam"
    if loc=="Kollu":
        loc="Kollumangudi"  
    month=request.args.get('month') 
        
    search_query = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'Name')
    sort_order = request.args.get('sort_order', None)
    # Apply search filtering
    # if search_query:
    #     total_items = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).count()
    #     if total_items == 0:
    #         flash("No results found for the search query.", "error")
    #         return redirect(url_for('Home', page=1, search=""))
    #     data = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).order_by(Employee.Name.asc()).paginate(page=page, per_page=selected_page_size)
    # else:
    #     total_items = Employee.query.count()
    #     data = Employee.query.order_by(Employee.Name.asc()).paginate(page=page, per_page=selected_page_size)

    if search_query:
        # Filter based on search query
        data = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%"))
        total_items = data.count()
        if total_items == 0:
            flash("No results found for the search query.", "error")
            return redirect(url_for('Home', page=1, search=""))
    else:
        # Get total items without search query
        data = Employee.query
        total_items = data.count()

    if project:
        data = data.filter(Employee.Project == project)

    if designation:
        data = data.filter(Employee.Designation == designation)

    if employment_status:
        data = data.filter(Employee.Employment_status == employment_status)

    if status:
        data = data.filter(Employee.employee_status == status)

    if loc:
        data = data.filter(Employee.Location == loc)  
    if month:
        
        # Extract the month part of the date
        data = data.filter(Employee.Joining_date != None )
        data = data.filter(func.substr(Employee.Joining_date, 4, 2) == month)
        
    if sort_order:
        if sort_order == 'asc':
            data = data.order_by(getattr(Employee, sort_by).asc())
        else:
            data = data.order_by(getattr(Employee, sort_by).desc())
    if selected_page_size == 'All':
        data = data.paginate(page=1, per_page=total_items)
    else:    
        data = data.paginate(page=page, per_page=selected_page_size)
    # Handle pagination
    if selected_page_size == 'All':
        new_page_count = 1
    else:
        new_page_count = total_items // selected_page_size
        if total_items % selected_page_size > 0:
            new_page_count += 1

    # Ensure the current page is within the valid range
    if page > new_page_count:
        page = new_page_count

    start_index = (page - 1) * selected_page_size if selected_page_size != 'All' else 0

    total_pages = data.pages if selected_page_size != 'All' else 1

    return render_template("index.html", data=data,
                           page_size_options=page_size_options,
                           selected_page_size=selected_page_size,
                           total_items=total_items, total_pages=total_pages,
                           start_index=start_index, search_query=search_query,page=page,project=project,designation=designation,employment_status=employment_status,status=status,loc=loc,month=month)
@app.route("/add", methods=["GET", "POST"])
def Add():
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts = joining_date.split('-')
        if len(date_parts) == 3:
            formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date, "%d-%m-%Y")
            current_date=date.today()   
            experience=current_date.year -int(join_date.year)
            if (current_date.month,current_date.day) < (join_date.month,join_date.day):
                experience-=1
            if experience <1:
                experience="Less than 1 year"       
        else:
            formatted_date = None 
            experience=None
        
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee_status=request.form.get("status")
        existing_data=Employee.query.filter_by(Name=name).first()
        if existing_data:
            flash(f'Employee {name} already exists!', 'error')
        if not existing_data:
            employee = Employee(
                Emp_id=emp_id,
                Name=name,
                Designation=designation,
                Department=department,
                Project=project,
                Job_role=job_role,
                Employment_status=employment_status,
                Joining_date=formatted_date,
                Experience=experience,
                Location=location,
                Last_promoted=last_promoted,
                Comments=comments,
                employee_status=employee_status
            )
            db.session.add(employee)
            db.session.commit()
            flash(f'Added {name} successfully!', 'success')
        return redirect(url_for('Add'))
    return render_template("add.html")

@app.route("/update/<int:sno>",methods=["GET","POST"])
def Update(sno):
    selected_date = request.args.get("date")
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts=joining_date.split("-")
        if len(date_parts)==3:
            formatted_date=f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date,"%d-%m-%Y")
            current_day=date.today()
            experience=current_day.year-int(join_date.year)
            if (current_day.month,current_day.day) < (join_date.month,join_date.day):
                experience-=1
            if experience < 1:
                experience="Less than 1 year"
        else:
            formatted_date = None 
            experience=None            
        
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee_status=request.form.get("status")
        employee=Employee.query.filter_by(Sno=sno).first()
        employee.Emp_id=emp_id
        employee.Name=name
        employee.Designation=designation
        employee.Department=department
        employee.Project=project
        employee.Job_role=job_role
        employee.Employment_status=employment_status
        employee.Joining_date=formatted_date
        employee.Experience=experience
        employee.Location=location
        employee.Last_promoted=last_promoted
        employee.Comments=comments
        employee.employee_status=employee_status
        db.session.add(employee)
        db.session.commit()
        flash(f'{name} updated successfully!', 'success')
        return redirect(url_for('Update', sno=sno))
    employee=Employee.query.filter_by(Sno=sno).first()
    return render_template("update.html",employee=employee,selected_date=selected_date)

@app.route("/delete/<int:sno>")
def Delete(sno):
    employee=Employee.query.filter_by(Sno=sno).first()
    delete=Delete_user(Name=employee.Name,Date=employee.Joining_date)
    db.session.add(delete)
    db.session.commit()
    db.session.delete(employee)
    db.session.commit()
    return redirect("/home")
with app.app_context():
        db.create_all()
        roles,leads,overall_counts,weekly_statistics = get_distinct_statistics()
        # print("leads",leads)
        # print("Distinct Roles:", roles)
        # data=extract_data_from_excel()

        

@app.route("/bulk",methods=["GET","POST"])
def bulk():
    if request.method=="POST":
        file = request.files['file']
        if file and allowed_file(file.filename):
            if file.filename.endswith(".xlsx"):
                try:
                    wb=load_workbook(file)
                    ws=wb.active
                    column_mappings = {
                        'Sno': 0,
                        'Emp_id': 1,
                        'Name': 2,
                        'Designation': 3,
                        'Department': 4,
                        'Project': 5,
                        'Job_role': 6,
                        'Employment_status': 7,
                        'Joining_date': 8,
                        'Experience': 9,
                        'Location': 10,
                        'Last_promoted': 11,
                        'Comments': 12
                        }
                    records_added = False
                    for row in ws.iter_rows (min_row=2,values_only=True):
                        if all(cell is None for cell in row):
                            continue
                        if not all(cell is None for cell in row):

                            Name = row[column_mappings['Name']]
                            if not Name:
                                continue
                            Emp_id = row[column_mappings['Emp_id']]                           
                            Designation = row[column_mappings['Designation']]
                            Department = row[column_mappings['Department']]
                            Project = row[column_mappings['Project']]
                            Job_role = row[column_mappings['Job_role']]
                            Employment_status = row[column_mappings['Employment_status']]
                            Joining_date = row[column_mappings['Joining_date']]
                            Experience = row[column_mappings['Experience']]
                            if Experience is None or Experience == "":
                                formatted_date = None
                                if isinstance(Joining_date, datetime):
                                    join_date = Joining_date
                                    formatted_date = join_date.strftime("%d-%m-%Y")
                                    month = join_date.month
                                    day = join_date.day
                                    year = join_date.year
                                    current_date = date.today()
                                    Experience = current_date.year - year
                                    if (current_date.month, current_date.day) < (month, day):
                                        Experience -= 1
                                    if Experience < 1:
                                        Experience = "Less than 1 year"
                                else:
                                    # If Joining_date is not a datetime object, set Experience to None
                                    Experience = None
                            else:
                                # If Experience is provided, keep formatted_date as is
                                formatted_date = Joining_date.strftime("%d-%m-%Y") if isinstance(Joining_date, datetime) else None
                    
                            Location = row[column_mappings['Location']]
                            Last_promoted = row[column_mappings['Last_promoted']]
                            Comments = row[column_mappings['Comments']]
                            employee_status="active"
                            existing_data=Employee.query.filter_by(Name=Name).first()
                            if existing_data:
                                flash( f'Employee {Name} already exists', 'error')
                            if not existing_data:
                                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                        Department=Department, Project=Project, Job_role=Job_role,
                                        Employment_status=Employment_status, Joining_date=formatted_date,
                                        Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                        Comments=Comments,employee_status=employee_status)
                                db.session.add(employee)
                                records_added = True
                                
                    db.session.commit()
                    if records_added:
                        flash('File uploaded successfully!', 'success')
                    
                    return redirect(url_for('Home'))            
                except Exception as e:
                    flash(f'Error: {e}', 'error')
                    return redirect(url_for('bulk'))
    
    return render_template("bulk.html")
    
@app.route("/view/<int:sno>")
def view(sno):
    data=Employee.query.filter_by(Sno=sno).first()
    return render_template("view.html",data=data)    
@app.route("/register",methods=["GET","POST"])
def register():
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        role = request.form["role"]
        name=request.form["name"]
        mobile=request.form["mobile"]
        filename=""
        existing_user = Login.query.filter_by(email=email).first()
        if existing_user:
            flash(f"{email} already exists!", "danger")
            return redirect(url_for("register"))
        if 'photo' in request.files:
            photo=request.files['photo']
            if photo.filename!='':
                filename=secure_filename(photo.filename)
                # print(filename)
                file_path=os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'],filename)
                photo.save(file_path)
        user=Login(email=email,password=password,Role=role,Name=name,MobileNumber=mobile,photo_filename=filename)
        db.session.add(user)
        db.session.commit()
        flash(f"Registration successful for {name}!", "success")
        return redirect(url_for("register"))

    return render_template("register.html")
@app.route("/get_employees_list/<employment_status>")
def get_employees_list(employment_status):
    employees=Employee.query.filter_by(Employment_status=employment_status).all()
    return render_template('employee_list.html', employees=employees,employment_status=employment_status)
    # employee_names=[employee.Name for employee in employees]
    # return jsonify({'employeeList': employee_names})

@app.route("/resume",methods=["GET","POST"])
def resume():
    flash_message=None
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    current_month = datetime.now().strftime("%B")
    
    if request.method=="POST":
        selected_tag=request.form['tag']
        month=request.form['month']
        files=request.files.getlist('file')
        sucessfully_uploaded=False
        for file in files:
            if file and allowed_files(file.filename):
                filename=secure_filename(file.filename)
                new_filename=f"{selected_tag}_{filename}"
                target_path=os.path.join(app.config['UPLOAD_FOLDER'],new_filename)
                if not os.path.exists(target_path):
                    file.save(target_path)
                    
                    resume=Resume(filename=new_filename,Month=month)
                    db.session.add(resume)
                    db.session.commit()
                    sucessfully_uploaded=True
                else:
                     flash(f'{filename} already exists !', 'error')
                    
        if sucessfully_uploaded:
            flash('Resume(s) uploaded successfully!', 'success')                
        return redirect(url_for('resume'))        
    return render_template("resume.html",months=months,current_month=current_month)

# @app.route("/employee_management", methods=["GET", "POST"])
# def employee():
#     default_page_size = 10
#     page_size_options = [10, 20, 30, 40, 50]

#     # Handle form submission for page size change (if applicable)
#     if request.method == "POST":
#         # Safely get the page size from form data and convert it
#         selected_page_size = request.form.get("page_size", default_page_size)
#         try:
#             selected_page_size = int(selected_page_size)
#         except ValueError:
#             selected_page_size = default_page_size
#         session['page_size'] = selected_page_size  # Update session variable

#     # Get data based on current page and retrieved/default page size
#     page = request.args.get('page', 1, type=int)
#     total_items = Resume.query.count()

#     # Retrieve selected page size from session (or default)
#     selected_page_size = session.get('page_size', default_page_size)
#     start_index = (page - 1) * selected_page_size
#     if selected_page_size != default_page_size:
#         new_page_count = total_items // selected_page_size
#         if total_items % selected_page_size > 0:
#             new_page_count += 1  # Account for partial last page
#         page = min(page, new_page_count)

#     resumes = Resume.query.paginate(page=page, per_page=selected_page_size)
#     no_of_pages = resumes.pages
    
#     return render_template("employee.html", resumes=resumes,
#                            page_size_options=page_size_options,
#                            selected_page_size=selected_page_size,
#                            total_items=total_items, no_of_pages=no_of_pages,start_index=start_index)

@app.route("/employee_management", methods=["GET", "POST"])
def employee():
    role=None
    email=session.get('email')
    if email:
        user=Login.query.filter_by(email=email).first()
        role=user.Role
        
        if not user:
            return redirect(url_for('login'))

    file_link = session.pop('file_link', None)
    public_key=os.environ.get('EMAILJS_PUBLIC_KEY')
    EMAILJS_SERVICE_ID=os.environ.get('EMAILJS_SERVICE_ID')
    EMAILJS_TEMPLATE_ID=os.environ.get('EMAILJS_TEMPLATE_ID_TABLE')
    # Prepare the email template parameters
    
    
    
    search_query = request.args.get("search", "").strip()
    
    # qa_lead_query = request.args.get('qa_lead', '').strip()
    filter_role = request.args.get('role') 
    filter_week=request.args.get('week')
    
    default_page_size = 10
    page_size_options = [10, 20, 30, 40, 50]
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    if request.method == "POST":
        selected_page_size = int(request.form.get("page_size", default_page_size))
        session['page_size'] = selected_page_size
        # Redirect to the same page with updated page size to avoid form resubmission issues
        return redirect(url_for('employee', page=1))
    else:
        selected_page_size = session.get('page_size', default_page_size)
    
    
    current_month = datetime.now().strftime("%B")
    selected_month = request.args.get("month", current_month)
    query = Resume.query

    if search_query:
        query = query.filter(
            or_(Resume.Name.ilike(f"%{search_query}%"), Resume.QA_Lead.ilike(f"%{search_query}%"))
            )
    
    if filter_role:  # Filter by role if provided
        query = query.filter(Resume.Role == filter_role)
    if filter_week:  # Filter by role if provided
        query = query.filter(Resume.week == filter_week)       
    if selected_month and not search_query :
        query = query.filter(Resume.Month == selected_month)
    page = request.args.get('page', 1, type=int)
    total_items =query.count()
    
       
    # Handle case where there are no items
    if total_items == 0:
        data =query.paginate(page=1, per_page=selected_page_size)
        start_index = 0
        end_index=0
        total_pages = 1
    else:
        # Handle pagination
        new_page_count = total_items // selected_page_size
        if total_items % selected_page_size > 0:
            new_page_count += 1

        # Ensure the current page is within the valid range
        if page > new_page_count:
            page = new_page_count

        start_index = (page - 1) * selected_page_size
        
        data = query.paginate(page=page, per_page=selected_page_size)
        if start_index:
            total_pages = 0
            end_index = 0
            
        else:    
            total_pages = data.pages
            end_index = min(start_index + selected_page_size, total_items)
    return render_template("employee.html", resumes=data,
                           page_size_options=page_size_options,
                           selected_page_size=selected_page_size,
                           total_items=total_items, total_pages=total_pages,
                           start_index=start_index,months=months,current_month=current_month,selected_month=selected_month,file_link=file_link,public_key=public_key,service_id=EMAILJS_SERVICE_ID,template_id=EMAILJS_TEMPLATE_ID,search_query=search_query,role=role,filter_role=filter_role,filter_week=filter_week,end_index=end_index)


@app.route("/view_resume/<filename>")
def view_resume(filename):
    resume_folder='static/files'
    file_path=os.path.join(resume_folder, filename)
    if not os.path.exists(file_path):
        abort(404)
    response=send_from_directory(resume_folder, filename)
    if filename.endswith('.docx'):
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        response.headers['Content-Disposition'] = f'inline; filename={filename}'
        # print(response.headers['Content-Type'],response.headers['Content-Disposition'])
    elif filename.endswith('.pdf'):
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename={filename}'
    else:
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
       
    return response



@app.route("/zip",methods=["GET","POST"])
def zip():
    current_month = datetime.now().strftime("%B")
    sucessful_message=None
    if request.method=="POST":
        
        zip_files=request.files.getlist('zipFiles')
        selected_tag=request.form['tag']
        for zip_file in zip_files:
            if zip_file and zip_file.filename.endswith('.zip'):
                temp_dir="temp_dir"
                os.makedirs(temp_dir,exist_ok=True)
                try:
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    for root,dirs,files in os.walk(temp_dir):
                        for filename in files:
                            file_path = os.path.join(root, filename)
                            if allowed_files(filename):
                                tag_filename=f"{selected_tag}_{filename}"
                                target_path=os.path.join(app.config['UPLOAD_FOLDER'], tag_filename)
                                # print("File")
                                if not os.path.exists(target_path):
                                    shutil.move(file_path,target_path)
                                    resume=Resume(filename=tag_filename,Month=current_month)
                                    db.session.add(resume)
                                    db.session.commit()
                                    sucessful_message=True
                                    
                                else:
                                    flash('Resume already exists!', 'error')
                                    
                finally:
                    shutil.rmtree(temp_dir)   

        if sucessful_message:
            flash('Resume(s) uploaded successfully!', 'success')

        return redirect(url_for('zip'))            
    return render_template("zip.html")

@app.route("/introcall/<int:resume_id>", methods=['GET', 'POST'])
def introCall(resume_id):
    selected_panel=""
    resume=Resume.query.get(resume_id)
    def safe_get(value):
        return value if value else "N/A"
    candidate_name=safe_get(resume.Name)
    candidate_email = safe_get(resume.Email)
    is_mandatory_missing = candidate_name == "N/A" or candidate_email == "N/A"
    candidate_phone = safe_get(resume.Phone)
    candidate_role = safe_get(resume.Role)
    candidate_experience = safe_get(resume.Experience)
    candidate_location = safe_get(resume.Location)
    candidate_notice_period = safe_get(resume.Notice_period)
    candidate_actual_ctc = safe_get(resume.Actual_CTC)
    candidate_expected_ctc = safe_get(resume.Expected_CTC)
    candidate_current_link = url_for('resume_details', resume_id=resume_id, _external=True)
    candidate_next_link = url_for('interview1v', resume_id=resume_id, _external=True)
    existing_entry=Intro.query.filter_by(resumeId=resume.id).first()
    all_panels = Panel.query.all()
    if existing_entry:
        status1 = existing_entry.Status
        comments1 = existing_entry.Comments
        selected_panel = existing_entry.SelectedPanel
        date = existing_entry.Date
        selected_panels = selected_panel.split(",") if selected_panel else []
        comments_dict = json.loads(existing_entry.Json_comments) if existing_entry.Json_comments else {}
    else:
        status1 = None
        comments1 = None
        selected_panel = ""
        date = None
        selected_panels = []
        comments_dict = {}
        
    if request.method == 'POST':
        
        
        date=request.form["date"]
        status=request.form["status"]
        comments=request.form["comments"]
        selected_panel=request.form["selectedPanel"]
        selected_panels = selected_panel.split(",") if selected_panel else []
        
        panel_comments = {panel: request.form.get(f"comment_{panel}", "") for panel in selected_panels}
        
        existing_entry=Intro.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
            existing_entry.Json_comments = json.dumps(panel_comments)
            status1=existing_entry.Status
            comments1=existing_entry.Comments
             
        else:    
            entry=Intro( Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel,Json_comments=json.dumps(panel_comments))
            db.session.add(entry)
            status1=status
            comments1=comments
            comments_dict = panel_comments

        db.session.commit()
        if status == "Rejected":
            flash("Candidate Rejected", "danger")  # Using 'danger' for a red flash message    
        elif status == "Move to Interview 1":
            interview_url = url_for('interview1v', resume_id=resume_id)
            flash(f"Candidate Moved to L1 — <a href='{interview_url}' class='alert-link'>Go to L1</a>", "success")
            
        return redirect(url_for('introCall', resume_id=resume.id))
    return render_template("intro.html",resume=resume,comments1=comments1,status1=status1,selected_panel=selected_panel,date=date,selected_panels=selected_panels,comments_dict=comments_dict,all_panels=all_panels,candidate_name=candidate_name,candidate_email = candidate_email,candidate_phone=candidate_phone,candidate_role=candidate_role,candidate_experience=candidate_experience,candidate_location=candidate_location,candidate_notice_period=candidate_notice_period,candidate_actual_ctc=candidate_actual_ctc,candidate_expected_ctc=candidate_expected_ctc,candidate_current_link=candidate_current_link,candidate_next_link=candidate_next_link,is_mandatory_missing=is_mandatory_missing)
@app.route("/interview1")
def interview1():
    return render_template("interview1.html")

@app.route("/interview2")
def interview2():
    return render_template("interview2.html")

@app.route("/introv",methods=["GET","POST"])
def introv():
    if request.method=="POST":
        return render_template("introv.html")

@app.route("/resume_details/<int:resume_id>")
def resume_details(resume_id):
    resume=Resume.query.get(resume_id)
    intro = Intro.query.filter_by(resumeId=resume_id).first()
    interview1 = Interview1.query.filter_by(resumeId=resume_id).first()
    interview2 = Interview2.query.filter_by(resumeId=resume_id).first()
    hr = Hr.query.filter_by(resumeId=resume_id).first()

    intro_status = intro.Status if intro else "Intro call not conducted"
    interview1_status = interview1.Status if interview1 else "Interview 1 not conducted"
    interview2_status = interview2.Status if interview2 else "Interview 2 not conducted"
    hr_status = hr.Status if hr else "HR not conducted"

    ## Panel feedbacks
    intro_panel_feedback = json.loads(intro.Json_comments) if intro and intro.Json_comments else {}
    interview1_panel_feedback = json.loads(interview1.Json_comments) if interview1 and interview1.Json_comments else {}
    interview2_panel_feedback = json.loads(interview2.Json_comments) if interview2 and interview2.Json_comments else {}
    hr_panel_feedback = json.loads(hr.Json_comments) if hr and hr.Json_comments else {}

    statuses = [intro_status, interview1_status, interview2_status, hr_status]
    all_rounds_status = "Cleared"

    for status in statuses:
        if status == "Rejected" or "not conducted" in status:
            all_rounds_status = "Rejected"
            break
    intro_call=Intro.query.filter_by(resumeId=resume_id).first()
    interview1=Interview1.query.filter_by(resumeId=resume_id).first()
    interview2=Interview2.query.filter_by(resumeId=resume_id).first()
    hr=Hr.query.filter_by(resumeId=resume_id).first()
    return render_template("resume_details.html",resume=resume,intro_call=intro_call,interview1=interview1,interview2=interview2,hr=hr,intro_status=intro_status,
        interview1_status=interview1_status,
        interview2_status=interview2_status,
        hr_status=hr_status,
        all_rounds_status=all_rounds_status,
        intro_panel_feedback=intro_panel_feedback,
        interview1_panel_feedback=interview1_panel_feedback,
        interview2_panel_feedback=interview2_panel_feedback,
        hr_panel_feedback=hr_panel_feedback
        )

# @app.route("/interview1v/<int:resume_id>", methods=["GET", "POST"])
# def interview1v(resume_id):
#     resume=Resume.query.get(resume_id)
#     existing_entry=Interview1.query.filter_by(resumeId=resume.id).first()
#     if existing_entry:
#         status1 = existing_entry.Status
#         comments1 = existing_entry.Comments
#         selected_panel=existing_entry.SelectedPanel
#         date = existing_entry.Date
#     else:
#         status1 = None
#         comments1 = None
#         selected_panel=""
#         date = None
#     if request.method=="POST":
#         date=request.form["date"]
#         comments=request.form["comments"]
#         status=request.form["status"]
        
#         selected_panel=request.form["selectedPanel"]
#         existing_entry=Interview1.query.filter_by(resumeId=resume.id).first()
#         if existing_entry:
#             existing_entry.Date = date
#             existing_entry.Status = status
#             existing_entry.Comments = comments
#             existing_entry.SelectedPanel = selected_panel
#             existing_entry.Status = status
#             existing_entry.Comments = comments
#             status1=existing_entry.Status
#             comments1=existing_entry.Comments
#         else:    
#             entry=Interview1(Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel)
#             db.session.add(entry)
#             status1=status
#             comments1=comments
#         db.session.commit()
#         if status == "Rejected":
#             flash("Candidate Rejected", "danger")  # Using 'danger' for a red flash message
#         elif status == "Move to Interview 2":
#             flash("Candidate Moved to Interview 2", "success")  # Using 'success' for a green flash message
#             return redirect(url_for('interview2v', resume_id=resume_id))
#         elif status == "hold":
#             flash("Candidate kept on hold", "warning")  # Using 'success' for a yellow flash message    

#         return redirect(url_for('interview1v', resume_id=resume.id))
#     resume = Resume.query.get(resume_id)
#     return render_template("interview1.html", resume=resume,status1=status1,comments1=comments1,selected_panel=selected_panel,date=date)

@app.route("/interview1v/<int:resume_id>", methods=["GET", "POST"])
def interview1v(resume_id):
    selected_panel=""
    resume=Resume.query.get(resume_id)
    def safe_get(value):
        return value if value else "N/A"
    candidate_name=safe_get(resume.Name)
    candidate_email = safe_get(resume.Email)
    is_mandatory_missing = candidate_name == "N/A" or candidate_email == "N/A"
    candidate_phone = safe_get(resume.Phone)
    candidate_role = safe_get(resume.Role)
    candidate_experience = safe_get(resume.Experience)
    candidate_location = safe_get(resume.Location)
    candidate_notice_period = safe_get(resume.Notice_period)
    candidate_actual_ctc = safe_get(resume.Actual_CTC)
    candidate_expected_ctc = safe_get(resume.Expected_CTC)
    candidate_current_link = url_for('resume_details', resume_id=resume_id, _external=True)
    candidate_next_link = url_for('interview2v', resume_id=resume_id, _external=True)
    existing_entry=Interview1.query.filter_by(resumeId=resume.id).first()
    all_panels = Panel.query.all()
    if existing_entry:
        status1 = existing_entry.Status
        comments1 = existing_entry.Comments
        selected_panel = existing_entry.SelectedPanel
        date = existing_entry.Date
        selected_panels = selected_panel.split(",") if selected_panel else []
        comments_dict = json.loads(existing_entry.Json_comments) if existing_entry.Json_comments else {}
    else:
        status1 = None
        comments1 = None
        selected_panel = ""
        date = None
        selected_panels = []
        comments_dict = {}
        
    if request.method == 'POST':
        
        
        date=request.form["date"]
        status=request.form["status"]
        comments=request.form["comments"]
        selected_panel=request.form["selectedPanel"]
        selected_panels = selected_panel.split(",") if selected_panel else []
        
        panel_comments = {panel: request.form.get(f"comment_{panel}", "") for panel in selected_panels}
        
        existing_entry=Interview1.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
            existing_entry.Json_comments = json.dumps(panel_comments)
            status1=existing_entry.Status
            comments1=existing_entry.Comments
             
        else:    
            entry=Interview1( Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel,Json_comments=json.dumps(panel_comments))
            db.session.add(entry)
            status1=status
            comments1=comments
            comments_dict = panel_comments

        db.session.commit()
        if status == "Rejected":
            flash("Candidate Rejected", "danger")  # Using 'danger' for a red flash message
        if status == "On Hold":
            flash("Candidate Kept on hold", "warning") 
        elif status == "Move to Interview 2":
            interview_url = url_for('interview2v', resume_id=resume_id)
            if g.user_role == 'admin':
                flash(f"Candidate Moved to L2 — <a href='{interview_url}' class='alert-link'>Go to L2</a>", "success")
        return redirect(url_for('interview1v', resume_id=resume.id))
    return render_template("interview1.html",resume=resume,comments1=comments1,status1=status1,selected_panel=selected_panel,date=date,selected_panels=selected_panels,comments_dict=comments_dict,all_panels=all_panels,candidate_name=candidate_name,candidate_email = candidate_email,candidate_phone=candidate_phone,candidate_role=candidate_role,candidate_experience=candidate_experience,candidate_location=candidate_location,candidate_notice_period=candidate_notice_period,candidate_actual_ctc=candidate_actual_ctc,candidate_expected_ctc=candidate_expected_ctc,candidate_current_link=candidate_current_link,candidate_next_link=candidate_next_link,is_mandatory_missing=is_mandatory_missing)

@app.route("/interview2v/<int:resume_id>", methods=["GET", "POST"])
def interview2v(resume_id):
    selected_panel=""
    resume=Resume.query.get(resume_id)
    def safe_get(value):
        return value if value else "N/A"
    candidate_name=safe_get(resume.Name)
    candidate_email = safe_get(resume.Email)
    is_mandatory_missing = candidate_name == "N/A" or candidate_email == "N/A"
    candidate_phone = safe_get(resume.Phone)
    candidate_role = safe_get(resume.Role)
    candidate_experience = safe_get(resume.Experience)
    candidate_location = safe_get(resume.Location)
    candidate_notice_period = safe_get(resume.Notice_period)
    candidate_actual_ctc = safe_get(resume.Actual_CTC)
    candidate_expected_ctc = safe_get(resume.Expected_CTC)
    candidate_current_link = url_for('resume_details', resume_id=resume_id, _external=True)
    candidate_next_link = url_for('hr', resume_id=resume_id, _external=True)
    existing_entry=Interview2.query.filter_by(resumeId=resume.id).first()
    all_panels = Panel.query.all()
    if existing_entry:
        status1 = existing_entry.Status
        comments1 = existing_entry.Comments
        selected_panel = existing_entry.SelectedPanel
        date = existing_entry.Date
        selected_panels = selected_panel.split(",") if selected_panel else []
        comments_dict = json.loads(existing_entry.Json_comments) if existing_entry.Json_comments else {}
    else:
        status1 = None
        comments1 = None
        selected_panel = ""
        date = None
        selected_panels = []
        comments_dict = {}
        
    if request.method == 'POST':
        
        
        date=request.form["date"]
        status=request.form["status"]
        comments=request.form["comments"]
        selected_panel=request.form["selectedPanel"]
        selected_panels = selected_panel.split(",") if selected_panel else []
        
        panel_comments = {panel: request.form.get(f"comment_{panel}", "") for panel in selected_panels}
        
        existing_entry=Interview2.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
            existing_entry.Json_comments = json.dumps(panel_comments)
            status1=existing_entry.Status
            comments1=existing_entry.Comments
             
        else:    
            entry=Interview2( Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel,Json_comments=json.dumps(panel_comments))
            db.session.add(entry)
            status1=status
            comments1=comments
            comments_dict = panel_comments

        db.session.commit()
        if status == "Rejected":
            flash("Candidate Rejected", "danger")  # Using 'danger' for a red flash message
        if status == "On Hold":
            flash("Candidate Kept on hold", "warning")     
        elif status == "Move to HR Round":
            interview_url = url_for('hr', resume_id=resume_id)
            if g.user_role == 'admin':
                flash(f"Candidate Moved to HR — <a href='{interview_url}' class='alert-link'>Go to HR</a>", "success")
        return redirect(url_for('interview2v', resume_id=resume.id))
    return render_template("interview2.html",resume=resume,comments1=comments1,status1=status1,selected_panel=selected_panel,date=date,selected_panels=selected_panels,comments_dict=comments_dict,all_panels=all_panels,candidate_name=candidate_name,candidate_email = candidate_email,candidate_phone=candidate_phone,candidate_role=candidate_role,candidate_experience=candidate_experience,candidate_location=candidate_location,candidate_notice_period=candidate_notice_period,candidate_actual_ctc=candidate_actual_ctc,candidate_expected_ctc=candidate_expected_ctc,candidate_current_link=candidate_current_link,candidate_next_link=candidate_next_link,is_mandatory_missing=is_mandatory_missing)

@app.route("/hr/<int:resume_id>", methods=["GET", "POST"])
def hr(resume_id):
    selected_panel = ""
    resume = Resume.query.get(resume_id)
    def safe_get(value):
        return value if value else "N/A"
    candidate_name=safe_get(resume.Name)
    candidate_email = safe_get(resume.Email)
    is_mandatory_missing = candidate_name == "N/A" or candidate_email == "N/A"
    candidate_phone = safe_get(resume.Phone)
    candidate_role = safe_get(resume.Role)
    candidate_experience = safe_get(resume.Experience)
    candidate_location = safe_get(resume.Location)
    candidate_notice_period = safe_get(resume.Notice_period)
    candidate_actual_ctc = safe_get(resume.Actual_CTC)
    candidate_expected_ctc = safe_get(resume.Expected_CTC)
    candidate_current_link = url_for('resume_details', resume_id=resume_id, _external=True) 
    candidate_next_link = url_for('hr', resume_id=resume_id, _external=True)
    existing_entry=Hr.query.filter_by(resumeId=resume.id).first()
    all_panels = Panel.query.all()

    if existing_entry:
        status1 = existing_entry.Status
        comments1 = existing_entry.Comments
        selected_panel = existing_entry.SelectedPanel
        date = existing_entry.Date
        selected_panels = selected_panel.split(",") if selected_panel else []
        comments_dict = json.loads(existing_entry.Json_comments) if existing_entry.Json_comments else {}
    else:
        status1 = None
        comments1 = None
        selected_panel = ""
        date = None
        selected_panels = []
        comments_dict = {}

    if request.method == "POST":
        date = request.form["date"]
        status = request.form["status"]
        comments = request.form["comments"]
        selected_panel = request.form["selectedPanel"]
        selected_panels = selected_panel.split(",") if selected_panel else []

        # Capture panel feedback comments
        panel_comments = {panel: request.form.get(f"comment_{panel}", "").strip() for panel in selected_panels}

        # **Validation: Ensure all fields are filled**
        missing_fields = []
        if not date:
            missing_fields.append("Date")
        if not status:
            missing_fields.append("Status")
        if not selected_panel:
            missing_fields.append("Panel Member")
        for panel, comment in panel_comments.items():
            if comment == "":
                missing_fields.append(f"{panel} feedback")

        # If any required field is missing, show flash message and prevent submission
        if missing_fields:
            flash(f"Please fill in all required fields: {', '.join(missing_fields)}", "danger")
            return redirect(url_for('hr', resume_id=resume.id))

        # **Insert or Update Interview 2 Data**
        existing_entry = Hr.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
            existing_entry.Json_comments = json.dumps(panel_comments)
            status1 = existing_entry.Status
            comments1 = existing_entry.Comments
        else:
            entry = Hr(
                Date=date,
                Status=status,
                Comments=comments,
                resumeId=resume.id,
                SelectedPanel=selected_panel,
                Json_comments=json.dumps(panel_comments)
            )
            db.session.add(entry)
            status1 = status
            comments1 = comments
            comments_dict = panel_comments

        db.session.commit()

        # **Redirect Based on Status**
        if status == "Rejected":
            flash("Candidate Rejected", "danger")
        elif status == "Move to HR Process":
            flash("Candidate Moved to HR Process", "success")
            return redirect(url_for('hr', resume_id=resume_id))
        elif status == "On Hold":
            flash("Candidate kept on hold", "warning")

        return redirect(url_for('hr', resume_id=resume.id))

    return render_template("hr.html",resume=resume,comments1=comments1,status1=status1,
                            selected_panel=selected_panel,date=date,
                            selected_panels=selected_panels,
                            comments_dict=comments_dict,all_panels=all_panels,
                            candidate_name=candidate_name,
                            candidate_email = candidate_email,candidate_phone=candidate_phone,
                            candidate_role=candidate_role,candidate_experience=candidate_experience,
                            candidate_location=candidate_location,
                            candidate_notice_period=candidate_notice_period,
                            candidate_actual_ctc=candidate_actual_ctc,
                            candidate_expected_ctc=candidate_expected_ctc,
                            candidate_current_link=candidate_current_link,
                            candidate_next_link=candidate_next_link,
                            is_mandatory_missing=is_mandatory_missing
                            )

@app.route("/get_interview_status/<int:resume_id>")
def get_intro_status(resume_id):
    intro=Intro.query.filter_by(resumeId=resume_id).first()
    interview1=Interview1.query.filter_by(resumeId=resume_id).first()
    interview2=Interview2.query.filter_by(resumeId=resume_id).first()
    hr=Hr.query.filter_by(resumeId=resume_id).first()
    intro_status=intro.Status if intro else "Intro call not conducted"
    interview1_status=interview1.Status if interview1 else "Interview 1 not conducted"
    interview2_status=interview2.Status if interview2 else "Interview 2 not conducted"
    hr_status=hr.Status if hr else "Hr not conducted"
    statuses = [intro_status, interview1_status, interview2_status, hr_status]
    all_rounds_status = "Cleared"
    
    for status in statuses:
        if status == "Rejected" or "not conducted" in status:
            all_rounds_status = "Rejected"
            break
            
    return jsonify({
        'intro_status': intro_status,
        'interview1_status': interview1_status,
        'interview2_status': interview2_status,
        'hr_status': hr_status,
        'all_rounds_status': all_rounds_status
    })

@app.route("/employee_data", methods=["GET", "POST"])
def employeeData():
    default_page_size = 20
    page_size_options = [20, 30, 40,'All']
    
    # Handle form submission for page size change
    if request.method == "POST":
        selected_page_size = request.form.get("page_size", default_page_size)
        if selected_page_size != 'All':
            selected_page_size = int(selected_page_size)
        session['page_size'] = selected_page_size
        # Redirect to the same page with updated page size to avoid form resubmission issues
        return redirect(url_for('Home', page=1, search=request.args.get('search', '')))
    else:
        selected_page_size = session.get('page_size', default_page_size)

    search_query = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort_by', 'Name')
    sort_order = request.args.get('sort_order', None)
    # Apply search filtering
    # if search_query:
    #     total_items = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).count()
    #     if total_items == 0:
    #         flash("No results found for the search query.", "error")
    #         return redirect(url_for('Home', page=1, search=""))
    #     data = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%")).order_by(Employee.Name.asc()).paginate(page=page, per_page=selected_page_size)
    # else:
    #     total_items = Employee.query.count()
    #     data = Employee.query.order_by(Employee.Name.asc()).paginate(page=page, per_page=selected_page_size)

    if search_query:
        # Filter based on search query
        data = Employee.query.filter(Employee.Name.ilike(f"%{search_query}%"))
        total_items = data.count()
        if total_items == 0:
            flash("No results found for the search query.", "error")
            return redirect(url_for('Home', page=1, search=""))
    else:
        # Get total items without search query
        data = Employee.query
        total_items = data.count()

    if sort_order:
        if sort_order == 'asc':
            data = data.order_by(getattr(Employee, sort_by).asc())
        else:
            data = data.order_by(getattr(Employee, sort_by).desc())
    if selected_page_size == 'All':
        data = data.paginate(page=1, per_page=total_items)
    else:    
        data = data.paginate(page=page, per_page=selected_page_size)
    # Handle pagination
    if selected_page_size == 'All':
        new_page_count = 1
    else:
        new_page_count = total_items // selected_page_size
        if total_items % selected_page_size > 0:
            new_page_count += 1

    # Ensure the current page is within the valid range
    if page > new_page_count:
        page = new_page_count

    start_index = (page - 1) * selected_page_size if selected_page_size != 'All' else 0

    total_pages = data.pages if selected_page_size != 'All' else 1

    return render_template("index.html", data=data,
                           page_size_options=page_size_options,
                           selected_page_size=selected_page_size,
                           total_items=total_items, total_pages=total_pages,
                           start_index=start_index, search_query=search_query,page=page)


@app.route("/profile")
def profile():
    email = session.get('email')
    
    if email:
        user=Login.query.filter_by(email=email).first()
        if not user:
            return redirect(url_for('login'))
        emailId=user.email
        password=user.password
        role=user.Role
        mobile=user.MobileNumber
        name=user.Name
        pic=session.get('picture')
        
        picture=pic if pic else ""
        # print(picture)

        filename = user.photo_filename
        return render_template("profile.html",emailId=emailId,password=password,role=role,name=name,mobile=mobile,filename=filename,picture=picture)

@app.route("/get_role")
def get_role():
    email = session.get('email')
    if email:
        user=Login.query.filter_by(email=email).first()
        role=user.Role
        return jsonify({'role': role})
    return  jsonify({'error': 'Email not in session'}), 400       
@app.route("/signout")
def signout():
    session.pop('email', None)   
    session.pop('picture',None) 
    return redirect(url_for('signPage'))
    


@app.route('/edit_config')
def edit_config():
    return render_template("config_editor.html")
    
    # subprocess.run(['notepad', 'static/config.js'])  
    # return 'File edited successfully'  

@app.route("/update_profile", methods=["POST"])
def update_profile():
    email = session.get('email')
    
    if email:
        user = Login.query.filter_by(email=email).first()
        
        changes=[]
        new_name=request.form["name"]
        if user.Name !=new_name:
            user.Name=new_name
            changes.append("Name")
        new_mobile = request.form["mobile"]    
        if user.MobileNumber !=new_mobile:
            user.MobileNumber=new_mobile
            changes.append("Mobile")
        new_role = request.form.get("role")
        if user.Role != new_role:
            user.Role = new_role
            changes.append("Role") 
        photo_updated = False
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo.filename != '':
                # Delete previous photo if it exists
                if user.photo_filename:
                    previous_photo_path = os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'], user.photo_filename)
                    if os.path.exists(previous_photo_path):
                        os.remove(previous_photo_path)
                # Save new photo
                filename = secure_filename(photo.filename)
                file_path = os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'], filename)
                photo.save(file_path)
                user.photo_filename = filename
                photo_updated = True
        db.session.commit()
        if photo_updated:
                changes.append("Profile Picture")
        if changes:
            flash('{} changed.'.format(', '.join(changes)), 'success')
        return redirect(url_for('profile'))
         
@app.route('/update_config', methods=['POST'])
def update_config():
    try:
        # Get the updated config data from the request
        updated_config = request.get_json()

        # Write the updated config data to the config.js file
        with open('static/config.js', 'w') as config_file:
            config_file.write(f'var config = {json.dumps(updated_config)};')

        return jsonify({'message': 'Config updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500   

@app.route('/sign-in')
def sign_in():
    flow.redirect_uri = REDIRECT_URI  
    authorization_url, state = flow.authorization_url()
    session['state'] = state
    
    return redirect(authorization_url) 

@app.route('/sign-up')
def sign_up():
    flow.redirect_uri = REDIRECT_URI  
    authorization_url, state = flow.authorization_url()
    session['state'] = state
    session['is_sign_up'] = True
    return redirect(authorization_url) 

@app.route('/google_sign_in')
def google_signin():
    flow.redirect_uri = REDIRECT_URI
    try:
        flow.fetch_token(authorization_response=request.url)
    except Exception as e:
        return redirect(url_for('signout'))
    
    
    credentials = flow.credentials
    
    request_session = requests.Request()
    time.sleep(1)
    try:
        id_info = id_token.verify_oauth2_token(
            id_token=credentials.id_token,
            request=request_session,
            audience=CLIENT_ID
        )

    except Exception as e:
        print("Error verifying token:", str(e))  # Debugging output
        return redirect(url_for('dashBoard'))
    
    # session['credentials'] = credentials_to_dict(credentials)
    # print(session['credentials'])
    email = id_info.get('email')
    name= id_info.get('name')
    picture = id_info.get('picture')
    
    if not email.endswith('@qaoncloud.com'):
        flash('You do not have acess', 'warning')
        return redirect(url_for('signPage'))
    
    
    user=Login.query.filter_by(email=email).first()

    if "is_sign_up" in session and session['is_sign_up']:
        session.pop('is_sign_up', None)
        if not user:
            session['temp_email']=email
            session['temp_name']=name
            session['temp_picture']=picture
            return redirect(url_for('complete_registration'))
        else:
            flash('Account already exists. Please sign in.', 'warning')
            
            return redirect(url_for('signPage'))
    
    else :
        if not user:
            flash('Account does not exist. Please sign up first.', 'warning')
            return redirect(url_for('signPage'))
        
    session['email']=email
    session['name']=name
    session['picture']=picture

    return redirect(url_for('dashBoard'))

@app.route('/complete_registration', methods=['GET', 'POST'])
def complete_registration():
    if request.method=='POST':
        role = request.form.get('role')
        mobile = request.form.get('mobile')
        name=request.form.get('name')
        email = session.get('temp_email')
        name = session.get('temp_name')
        picture = session.get('temp_picture')

        if not email:
            return redirect(url_for('sign_up'))
        
        user = Login(email=email, password="", Role=role, Name=name, MobileNumber=mobile, photo_filename="")
        db.session.add(user)
        db.session.commit()

        session.pop('temp_email', None)
        session.pop('temp_name', None)
        session.pop('temp_picture', None)

        session['email'] = email
        session['name'] = name
        session['picture'] = picture

        return redirect(url_for('dashBoard'))
    return render_template('complete_registration.html')

@app.route("/delete_resume/<int:resume_id>", methods=["POST"])
def delete_resume(resume_id):
    selected_month = request.args.get('month')
    resume = Resume.query.get_or_404(resume_id)
    resume_folder = 'static/files'
    file_path = os.path.join(resume_folder, resume.filename)

    # Delete the resume entry from the database
    db.session.delete(resume)
    db.session.commit()

    # Check if the file exists and delete it
    if os.path.exists(file_path):
        os.remove(file_path)

    return redirect(url_for('employee', month=selected_month))

@app.route("/delete_selected_resumes", methods=["POST"])
def delete_selected_resumes():
    data = request.get_json()
    resume_ids = data.get("resume_ids", [])
    selected_month = request.args.get('month')
    
    if resume_ids:
        resumes = Resume.query.filter(Resume.id.in_(resume_ids)).all()
        for resume in resumes:
            # Delete the file
            resume_folder = 'static/files'
            file_path = os.path.join(resume_folder, resume.filename)
            if os.path.exists(file_path):
                os.remove(file_path)

            # Delete the database entry
            db.session.delete(resume)

        db.session.commit()

    return jsonify({"success": True, "message": "Selected resumes deleted!"}), 200

@app.route('/dmax_upload', methods=['GET', 'POST'])
def dmax_upload():
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    current_month = datetime.now().strftime("%B")
    if request.method == 'POST':
        file = request.files['file']
        selected_month = request.form.get('month')
        if file and file.filename.endswith('.xlsx'):
            wb = load_workbook(file)
            sheet_names = wb.sheetnames
            
            matching_sheet = None
            for sheet_name in sheet_names:
                if sheet_name.lower().startswith(selected_month.lower()):
                    matching_sheet = sheet_name
                    break
            if matching_sheet:
                ws = wb[matching_sheet]
            else:
                flash(f"No sheet found for the selected month: {selected_month}!", "error")
                return render_template('dmax_upload.html')    
            
            records_added = False

            for row in ws.iter_rows(min_row=3, values_only=True):
                if all(cell is None for cell in row):
                    continue
                designation = row[4]
                if designation:
                    designation = designation.strip().lower()
                    if designation=="intern":
                        process_and_insert_data(row, designation)
                    if designation=="jr. qa engineer" or designation=="qa automation engineer":
                        process_and_insert_data(row, designation)
                    if designation=="sr. qa engineer":
                        
                        process_and_insert_data(row, designation)
                    if designation=="qa engineer":
                        process_and_insert_data(row, designation) 
                    if designation == 'adm' or designation== 'qa lead':
                        process_and_insert_data(row, designation)  
            
                
            flash("Data successfully uploaded!", "success")
                    
            render_template('dmax_upload.html',months=months,current_month=current_month)
            

    return render_template('dmax_upload.html')

@app.route('/dmax_view')
def dmax_view():

    
    dmax_view_url = url_for('dmax_view')
    role = request.args.get('role') or ''
    location = request.args.get('location') or ''
    designation=request.args.get('designation') or''
    project=request.args.get('project') or ''
    month=request.args.get('month') or ''
    
    if location == "" or location is None:
        location = None
    if role== "" or role is None:
        role=None    
    if designation =="" or designation is None:
        designation=None   
    if project =="" or project is None:
        project=None
    if designation=="Jr QA Engineer":
        designation="Jr. QA Engineer"  
    if designation=="QA Engineer":
        designation="QA Engineer" 
    if designation=="Sr. QA Engineer":
        
        designation="Sr. QA Engineer"               
    
    # Query all records
    data = Dmax_tl.query
    intern_data=Dmax_intern.query
    jr_qa_data=Dmax_jr_qa_eng.query
    qa_data=Dmax_qa_eng.query
    sr_qa_data=Dmax_sr_qa_eng.query
    if role:
        if role == 'intern':
            if location:
                intern_data = intern_data.filter_by(Centre=location)
            if project:
                intern_data = intern_data.filter_by(Project=project)    
            if designation:
                intern_data = intern_data.filter_by(Designation=designation)
            if month:
                intern_data=intern_data.filter_by(Month=month)
            intern_data=intern_data.all()    
            # Set other roles' data to empty lists since we are only showing interns
            data=jr_qa_data = qa_data = sr_qa_data = []
        elif role == 'jr_qa_eng':
            
            
            if designation:
                jr_qa_data = jr_qa_data.filter_by(Designation=designation)
            if project:
                jr_qa_data = jr_qa_data.filter_by(Project=project)    
            if location:
                jr_qa_data = jr_qa_data.filter_by(Centre=location)
            if month:
                jr_qa_data=jr_qa_data.filter_by(Month=month)
            jr_qa_data = jr_qa_data.all()
            # Set other roles' data to empty lists
            data=intern_data = qa_data = sr_qa_data = []
        elif role == 'qa_eng':
            if designation:
                qa_data = qa_data.filter_by(Designation=designation)
                # print("h")
                
            if location:
                qa_data = qa_data.filter_by(Centre=location)
            if project:
                qa_data = qa_data.filter_by(Project=project)    
            if month:
                qa_data=qa_data.filter_by(Month=month)
            
            qa_data = qa_data.all()
            # Set other roles' data to empty lists
            data=intern_data = jr_qa_data = sr_qa_data = []
        elif role == 'sr_qa_eng':
            if location:
                sr_qa_data = sr_qa_data.filter_by(Centre=location)
            if project:
                sr_qa_data = sr_qa_data.filter_by(Project=project)    
            if designation:
                sr_qa_data = sr_qa_data.filter_by(Designation=designation)
            if month:
                sr_qa_data=sr_qa_data.filter_by(Month=month)
            sr_qa_data = sr_qa_data.all()
            # Set other roles' data to empty lists
            data=intern_data = jr_qa_data = qa_data = []
        elif role == 'team_lead':
            if location:
                data=data.filter_by(Centre=location)
            if project:
                data = data.filter_by(Project=project)    
            if designation:
                data=data.filter_by(Designation=designation) 
            if month:
                data=data.filter_by(Month=month)     
            data=data.all()      
            # Set other roles' data to empty lists
            intern_data = jr_qa_data = qa_data = sr_qa_data = []
    return render_template('dmax_view.html', data=data,intern_data=intern_data,jr_qa_data=jr_qa_data,qa_data=qa_data,sr_qa_data=sr_qa_data,location=location,role=role,project=project,designation=designation,month=month,dmax_view_url=dmax_view_url)
    
@app.route('/dmax_add', methods=['GET', 'POST'])
def dmax_add():
    return render_template("dmax_add.html")

@app.route("/dmax_add_teamlead", methods=['GET', 'POST'])
def Teamlead():
    if request.method == 'POST':
        emp_name = request.form.get('name')
        emp_id = request.form.get('emp_id')
        centre = request.form.get('location')
        designation = request.form.get('designation')
        project = request.form.get('project')
        month = request.form.get('month')
        target = request.form.get('target')
        actual = request.form.get('actual')
        production = request.form.get('production')
        quality = request.form.get('quality')
        location=request.form.get('location')          
        attrition=request.form.get('attrition')
        skill = request.form.get('skill')
        Dmaxsharing=request.form.get('dsharing')
        overall_dmax = request.form.get('odsharing')
        existing_employee = Dmax_tl.query.filter_by(EmployeeName=emp_name).first()
        if existing_employee:
                    flash(f"Employee {emp_name} already exists!", "error")
                    return redirect('/dmax_add')
        else:
            # If employee doesn't exist, create a new record
                new_intern = Dmax_tl(
                        Centre=location,
                        EmployeeName=emp_name,
                        EmpID=emp_id,
                        Designation=designation,
                        Project=project,
                        Month=month,
                        Target=target,
                        Actual=actual,
                        Attrition=attrition,
                        Production=production,
                        
                        Quality=quality,
                        Skill=skill,
                        DmaxSharing=Dmaxsharing,
                        OverallDmaxScore=overall_dmax 
                    )
                    
                    
                db.session.add(new_intern)
                db.session.commit()
                flash("Employee added successfully!", "success")
    return render_template("dmax_add_teamlead.html")  

@app.route("/dmax_add_intern", methods=['GET', 'POST'])
def Intern():
    if request.method == 'POST':
        emp_name = request.form.get('name')
        emp_id = request.form.get('emp_id')
        centre = request.form.get('location')
        designation = request.form.get('designation')
        project = request.form.get('project')
        month = request.form.get('month')
        target = request.form.get('target')
        actual = request.form.get('actual')
        production = request.form.get('production')
        quality = request.form.get('quality')
        location=request.form.get('location')        
        attendance=request.form.get('attendance')
        skill = request.form.get('skill')
        
        overall_dmax = request.form.get('odsharing')
        existing_employee = Dmax_intern.query.filter_by(EmployeeName=emp_name).first()
        if existing_employee:
                    flash(f"Employee {emp_name} already exists!", "error")
                    return redirect('/dmax_add')
        else:
            # If employee doesn't exist, create a new record
                new_intern = Dmax_intern(
                        Centre=location,
                        EmployeeName=emp_name,
                        EmpID=emp_id,
                        Designation=designation,
                        Project=project,
                        Month=month,
                        Target=target,
                        Actual=actual,
                        Attendance=attendance,
                        Production=production,
                        Quality=quality,
                        Skill=skill,
                        
                        OverallDmaxScore=overall_dmax 
                    )
                    
                    
                db.session.add(new_intern)
                db.session.commit()
                flash("Employee added successfully!", "success")
    return render_template("dmax_add_intern.html")  


@app.route("/dmax_add_jrqaeng", methods=['GET', 'POST'])
def jrqa():
    if request.method == 'POST':
        emp_name = request.form.get('name')
        emp_id = request.form.get('emp_id')
        centre = request.form.get('location')
        designation = request.form.get('designation')
        project = request.form.get('project')
        month = request.form.get('month')
        target = request.form.get('target')
        actual = request.form.get('actual')
        production = request.form.get('production')
        quality = request.form.get('quality')
        location=request.form.get('location')        
        attendance=request.form.get('attendance')
        skill = request.form.get('skill')
        
        overall_dmax = request.form.get('odsharing')
        existing_employee = Dmax_jr_qa_eng.query.filter_by(EmployeeName=emp_name).first()
        if existing_employee:
                    flash(f"Employee {emp_name} already exists!", "error")
                    return redirect('/dmax_add')
        else:
            # If employee doesn't exist, create a new record
                new_intern = Dmax_jr_qa_eng(
                        Centre=location,
                        EmployeeName=emp_name,
                        EmpID=emp_id,
                        Designation=designation,
                        Project=project,
                        Month=month,
                        Target=target,
                        Actual=actual,
                        Attendance=attendance,
                        Production=production,
                        Quality=quality,
                        Skill=skill,
                        OverallDmaxScore=overall_dmax 
                    )
                    
                    
                db.session.add(new_intern)
                db.session.commit()
                flash("Employee added successfully!", "success")
    return render_template("dmax_add_jrqa.html")

@app.route("/dmax_add_qaeng", methods=['GET', 'POST'])
def qaeng():
    if request.method == 'POST':
        emp_name = request.form.get('name')
        emp_id = request.form.get('emp_id')
        centre = request.form.get('location')
        designation = request.form.get('designation')
        project = request.form.get('project')
        month = request.form.get('month')
        target = request.form.get('target')
        actual = request.form.get('actual')
        production = request.form.get('production')
        quality = request.form.get('quality')
        location=request.form.get('location') 
        new_initiatives=request.form.get("")       
        attendance=request.form.get('attendance')
        skill = request.form.get('skill')
        
        overall_dmax = request.form.get('odsharing')
        existing_employee = Dmax_qa_eng.query.filter_by(EmployeeName=emp_name).first()
        if existing_employee:
                    flash(f"Employee {emp_name} already exists!", "error")
                    return redirect('/dmax_add')
        else:
            # If employee doesn't exist, create a new record
                new_intern = Dmax_qa_eng(
                        Centre=location,
                        EmployeeName=emp_name,
                        EmpID=emp_id,
                        Designation=designation,
                        Project=project,
                        Month=month,
                        Target=target,
                        Actual=actual,
                        Attendance=attendance,
                        Production=production,
                        New_initiatives=new_initiatives,
                        Quality=quality,
                        Skill=skill,
                        OverallDmaxScore=overall_dmax 
                    )
                    
                    
                db.session.add(new_intern)
                db.session.commit()
                flash("Employee added successfully!", "success")
    
    return render_template("dmax_add_qa.html")


@app.route("/dmax_add_srqaeng", methods=['GET', 'POST'])
def srqaeng():
    if request.method == 'POST':
        emp_name = request.form.get('name')
        emp_id = request.form.get('emp_id')
        centre = request.form.get('location')
        designation = request.form.get('designation')
        project = request.form.get('project')
        month = request.form.get('month')
        target = request.form.get('target')
        actual = request.form.get('actual')
        production = request.form.get('production')
        quality = request.form.get('quality')
        location=request.form.get('location') 
        new_initiatives=request.form.get("")       
        attendance=request.form.get('attendance')
        skill = request.form.get('skill')
        
        overall_dmax = request.form.get('odsharing')
        existing_employee = Dmax_sr_qa_eng.query.filter_by(EmployeeName=emp_name).first()
        if existing_employee:
                    flash(f"Employee {emp_name} already exists!", "error")
                    return redirect('/dmax_add')
        else:
            # If employee doesn't exist, create a new record
                new_intern = Dmax_sr_qa_eng(
                        Centre=location,
                        EmployeeName=emp_name,
                        EmpID=emp_id,
                        Designation=designation,
                        Project=project,
                        Month=month,
                        Target=target,
                        Actual=actual,
                        Attendance=attendance,
                        Production=production,
                        New_initiatives=new_initiatives,
                        Quality=quality,
                        Skill=skill,
                        OverallDmaxScore=overall_dmax 
                    )
                    
                    
                db.session.add(new_intern)
                db.session.commit()
                flash("Employee added successfully!", "success")
    
    return render_template("dmax_add_sr_qa.html")

@app.route("/qa_requirements", methods=['GET', 'POST'])
def qareq():
    public_key=os.environ.get('EMAILJS_PUBLIC_KEY')
    service_id=os.environ.get('EMAILJS_SERVICE_ID')
    template_id=os.environ.get('EMAILJS_TEMPLATE_ID')
    email = session.get('email')  
            
    if request.method == 'POST':
        
        target = request.form.get('target_date')
        role=request.form.get('role')
        Assign=request.form.get('assign')
        Req=request.form.get('req')
        
    return render_template("qareq.html",public_key=public_key,service_id=service_id,template_id=template_id,email=email)
 
@app.route("/excel_resume", methods=['GET', 'POST'])
def excel_resume():
    if request.method=="POST":
        current_month = datetime.now().strftime("%B")
        month=request.form['month_excel']
        print(month)
        if 'file' not in request.files:
            flash('No file part')
            
            return redirect(request.url)
        file = request.files['file']
    
        if file.filename == '':
            flash('No selected file')
            
            return redirect(request.url)
        if file:
            try:
                wb=load_workbook(file)
                
                sheet=wb.active
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                
                   if not all(cell is None for cell in row):
                        email=row[1]
                        name=row[2]
                        qualification=row[3]
                        phone=str(row[4])
                        location=row[5]
                        experience=row[6]
                        
                        lead=row[7]
                        result=row[8]
                        current_ctc=row[9]
                        expecting_ctc=row[10]
                        notice_period=row[11]
                        suggestions=row[12]
                        link=row[13]
                        role = map_role_based_on_experience(experience)
                        
                        # print(result,current_ctc,expecting_ctc,notice_period,suggestions)
                        try:
                            # Convert the experience to a string, strip spaces, and remove any non-numeric characters (like 'yrs')
                            experience = str(experience).strip() if experience else None
                            
                            
                            if experience:
                                
                                # Remove non-numeric characters like "yrs" using a regular expression
                                # This keeps only digits and a decimal point
                                numeric_experience = re.sub(r'[^\d.]', '', experience)
                                
                                # Try converting the cleaned value to a float
                                if numeric_experience:
                                    float_experience = float(numeric_experience)
                                    experience = str(float_experience)  # Convert back to string if successful
                                    print("yrs experience",experience)  # Print for debugging
                                else:
                                    experience = None  # If cleaning results in an empty string, set to None
                            else:
                                experience = None  # Set to None if empty
                        except Exception as e:
                            # Log error if needed, and set experience to None in case of failure
                            print(f"Error processing experience: {e}")
                            experience = None
                        existing_employee =Resume.query.filter_by(Email=email).first()
                        if not existing_employee:
                            
                            new_candidate = Resume(
                                Email=email,
                                filename=name,
                                Name=name,
                                Qualification=qualification,
                                Phone=phone,
                                Location=location,
                                Experience=experience,
                                QA_Lead=lead,
                                Link=link,
                                Role=role,
                                Actual_CTC=current_ctc,
                                Expected_CTC=expecting_ctc,
                                Notice_period=notice_period,
                                Month=month
                            )
                            
                            db.session.add(new_candidate)
                            db.session.commit() 
                            flash('Resume(s) uploaded successfully!', 'success')
                        else:
                            flash('Resume(s) already exists', 'error') 
                            
            except:
                pass    
    return redirect(url_for('resume'))  



# @app.route("/file")
# def file():
#     EMAILJS_USER_ID=os.environ.get('EMAILJS_PUBLIC_KEY')
#     public_key=os.environ.get('EMAILJS_PUBLIC_KEY')
#     EMAILJS_SERVICE_ID=os.environ.get('EMAILJS_SERVICE_ID_TABLE')
#     EMAILJS_TEMPLATE_ID=os.environ.get('EMAILJS_TEMPLATE_ID_TABLE')
#     print(EMAILJS_SERVICE_ID,EMAILJS_TEMPLATE_ID)
#     #email.js
#     accepted_interviews = db.session.query(Interview1.resumeId).filter_by(Status='Move to Interview 2').all()
#     accepted_resume_ids = [interview.resumeId for interview in accepted_interviews]

#     # Fetch resumes based on accepted interview IDs
#     resumes = db.session.query(Resume).filter(Resume.id.in_(accepted_resume_ids)).all()
#     resume_details = []

#     for resume in resumes:
#         # Extracting relevant fields dynamically from the resume object
#         details = {
#             'Name': resume.Name,
#             'Role': resume.Role,
#             'Experience': resume.Experience,
#             'Location': resume.Location,
#             'Month': resume.Month,
#             'Actual CTC': resume.Actual_CTC,
#             'Expected CTC': resume.Expected_CTC,
#             'Phone': resume.Phone,
#             'Email': resume.Email
#         }
#         resume_details.append(details)
#         print(resume_details)
        
#     resume_table_html = [
#         {
#             'Name': resume.Name,
#             'Role': resume.Role,
#             'Experience': resume.Experience,
#             'Location': resume.Location,
#             'Actual_CTC': resume.Actual_CTC,
#             'Expected_CTC': resume.Expected_CTC,
#             'Phone': resume.Phone,
#             'Email': resume.Email
#         } for resume in resumes
#     ]
#     print(resume_table_html)
#     if 'credentials' in session:
#         credentials = Credentials(
#         token=session['credentials']['token'],
#         refresh_token=session['credentials']['refresh_token'],
#         token_uri=session['credentials']['token_uri'],
#         client_id=session['credentials']['client_id'],
#         client_secret=session['credentials']['client_secret'],
#         scopes=session['credentials']['scopes']
#         )
#         try:
#             drive_service = build('drive', 'v3', credentials=credentials)
            
#             # Call your upload function or other operations here
#             # Example: upload_to_drive(drive_service, excel_file_path)
#         except Exception as e:
#             print("Error building Drive service:", e)
#             return "Error accessing Google Drive"
#     excel_file = create_excel_in_memory(resume_table_html)
#     return render_template("file.html",public_key=public_key)     

@app.route('/google_drive_auth')
def google_drive_auth():
    credentials = session.get('credentials')
    # print("Stored Credentials:", credentials) 
    if credentials:
        # Load existing credentials
        creds = Credentials(**credentials)
        # print("credentials scopes",creds.scopes)
        # Check if the required scope is included
        
            # If the necessary scope is not present, redirect to authorize again
        # print("Requesting new authorization for Drive scope.")
        drive_flow.redirect_uri = REDIRECT_URI_DRIVE
        authorization_url, state = drive_flow.authorization_url(prompt='consent')  # Request consent for new scope
        session['state'] = state
        return redirect(authorization_url)

    else:
        # No credentials found, initiate the OAuth process
        flow = drive_flow
        flow.redirect_uri = REDIRECT_URI_DRIVE
        authorization_url, state = flow.authorization_url(prompt='consent')
        session['state'] = state
        return redirect(authorization_url)    

    return None

@app.route('/google_drive_callback')
def google_drive_callback():
    flow = drive_flow  # Use the same flow instance
    flow.redirect_uri = REDIRECT_URI_DRIVE
    authorization_response = request.url
    credentials = flow.fetch_token(authorization_response=authorization_response)
    
    session['credentials'] = credentials_to_dict(credentials)  # Store credentials in session
    # print("check this",session['credentials'])
    return redirect(url_for('upload_to_drive'))

    

@app.route('/upload_to_drive')
def upload_to_drive():
    public_key=os.environ.get('EMAILJS_PUBLIC_KEY')
    EMAILJS_SERVICE_ID=os.environ.get('EMAILJS_SERVICE_ID')
    EMAILJS_TEMPLATE_ID=os.environ.get('EMAILJS_TEMPLATE_ID_TABLE')
    try:
        credentials = session.get('credentials')
        if not credentials:
            flash('No credentials found. Please authenticate first.', 'danger')
            return redirect(url_for('google_drive_auth'))
        accepted_interviews = db.session.query(Interview1.resumeId).filter_by(Status='Move to Interview 2').all()
        accepted_resume_ids = [interview.resumeId for interview in accepted_interviews]

        # Fetch resumes based on accepted interview IDs
        resumes = db.session.query(Resume).filter(Resume.id.in_(accepted_resume_ids)).all()
        resume_details = []

        # for resume in resumes:
        #     # Extracting relevant fields dynamically from the resume object
        #     details = {
        #         'Name': resume.Name,
        #         'Role': resume.Role,
        #         'Experience': resume.Experience,
        #         'Location': resume.Location,
        #         'Month': resume.Month,
        #         'Actual CTC': resume.Actual_CTC,
        #         'Expected CTC': resume.Expected_CTC,
        #         'Phone': resume.Phone,
        #         'Email': resume.Email
        #     }
        #     resume_details.append(details)
        #     print(resume_details)
            
        resume_data = [
            {
                'Name': resume.Name,
                'Role': resume.Role,
                'Experience': resume.Experience,
                'Location': resume.Location,
                'Actual_CTC': resume.Actual_CTC,
                'Expected_CTC': resume.Expected_CTC,
                'Phone': resume.Phone,
                'Email': resume.Email
            } for resume in resumes
        ]
        # Load credentials
        creds = Credentials(**credentials)
        # print("this is ",creds.scopes)
        # Create a Google Drive service object
        drive_service = build('drive', 'v3', credentials=creds)

        # Specify the path to your Excel file
        excel_file = create_excel_in_memory(resume_data) 
        file_metadata = {
            'name': 'resume.xlsx',
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # Just specify the file name
        }
        

        # Prepare the file for upload
        media = MediaIoBaseUpload(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Create the file in Google Drive
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        
        file_link = f'https://docs.google.com/spreadsheets/d/{file_id}/edit'

        # Make the file publicly accessible
        permission = {
            'type': 'anyone',
            'role': 'reader'
        }
        drive_service.permissions().create(fileId=file_id, body=permission).execute()
        session['file_link'] = file_link
        send_email_with_file_link(file_link,public_key,EMAILJS_SERVICE_ID,EMAILJS_TEMPLATE_ID)
        flash('File uploaded successfully!', 'success')
        
        return redirect(url_for('employee'))
    except Exception as e:
        # print(f'An error occurred: {str(e)}', 'danger')
        return redirect(url_for('employee'))   

@app.route('/edit_employee_resume/<int:employee_id>', methods=['GET', 'POST'])  
def edit_employee_resume(employee_id):
    # Fetch employee details from the database using employee_id
    resume = Resume.query.get(employee_id)
    ## pass the panel members to the template
    panel_members = Panel.query.all()
    if request.method=="POST":
        resume.Email=request.form['email']
        resume.Name=request.form['name']
        resume.Qualification=request.form['qualification']
        resume.Phone=request.form['phone']
        resume.Location=request.form['location']
        resume.Actual_CTC = request.form['actual_ctc']
        resume.Expected_CTC = request.form['expected_ctc']
        resume.QA_Lead=request.form['qa_lead']
        resume.Experience=request.form['experience']
        resume.Notice_period=request.form['notice_period']
        resume.Link=request.form["resume_link"]
        resume.Role=map_role_based_on_experience(resume.Experience)
        resume.week = request.form['week']
        
        db.session.commit()
        flash('Successfully updated!', 'success')
    return render_template("update_resume.html",resume=resume,panel_members=panel_members)


@app.route("/form_test", methods=["GET","POST"])
def form_test():
    if request.method == "POST":
        name = request.form.get("selectedPanel")
        email = request.form.get("panelEmail")
        return f"Received Name: {name}, Email: {email}"
    return render_template("form_test.html")


@app.route('/add_panel_member', methods=['POST'])
def add_panel_member():
    data = request.json
    name = data.get('name')
    email = data.get('email')
    
    existing_member = Panel.query.filter_by(name=name).first()
    if existing_member:
        return jsonify({"error": "Panel member already exists!"}), 409
    new_member = Panel(name=name, email=email)
    db.session.add(new_member)
    db.session.commit()

    return jsonify({"message": "Panel member added successfully!", "name": name, "email": email}), 201

    # data = request.json
    # name = data.get('name')
    # email = data.get('email')

    # if not name or not email:
    #     return jsonify({"error": "Name and email are required!"}), 400

    # # Check if the panel member already exists
    # existing_member = PanelMember.query.filter_by(name=name).first()
    # if existing_member:
    #     return jsonify({"error": "Panel member already exists!"}), 409

    # # Add new panel member
    # new_member = PanelMember(name=name, email=email)
    # db.session.add(new_member)
    # db.session.commit()

    # return jsonify({"message": "Panel member added successfully!", "name": name, "email": email}), 201


@app.route("/delete_panel", methods=["POST"])
def delete_panel():
    data = request.get_json()
    panel_name = data.get("panel")

    if not panel_name:
        return jsonify({"success": False, "error": "Invalid panel name"}), 400

    panel = Panel.query.filter_by(name=panel_name).first()
    if panel:
        db.session.delete(panel)
        db.session.commit()
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "Panel not found"}), 404  

if __name__ == "__main__":
    
    app.run(debug=True)



